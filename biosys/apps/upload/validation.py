import logging
import copy
import datetime

from openpyxl import load_workbook
from django.conf import settings

from main.models import SiteCharacteristic, SiteVisit, SiteVisitDataFile, Visit, Site
from species.models import Species
from vegetation.models import *
from animals.models import *
from main import utils as util_model
import utils_openpyxl as util_xls
from models import SiteVisitDataFileError
from utils import is_blank
from upload.validator_functions import ground_cover_validate, lat_lon_validate

logger = logging.getLogger(__name__)

ERRORS = [
    ("missing_data_datasheet", "The data sheet {sheet_name} is missing", logging.FATAL),
    ("no_data_for_model", "No data found for {model}", logging.FATAL),
    ("too_many_data_for_model", "Too many rows for {model}. It should only contains one row (or column)",
     logging.FATAL),
    ("extra_validation_error", "Row Validation Error (sheet:'{datasheet}', row:{row}) - {msg}", logging.ERROR),
    ("field_error", "Field error: (sheet:'{datasheet}', col:'{col}', row:{row}) - {msg}", logging.ERROR),
]


class ErrorManager():
    all_errors = None

    @staticmethod
    def parse_errors():
        return [Error(*t) for t in ERRORS]

    @staticmethod
    def all():
        if ErrorManager.all_errors is None:
            ErrorManager.all_errors = ErrorManager.parse_errors()
        return ErrorManager.all_errors

    @staticmethod
    def first():
        return ErrorManager._clone(ErrorManager.all()[0])

    @staticmethod
    def last():
        return ErrorManager._clone(ErrorManager.all()[-1])

    @staticmethod
    def _first_or_none(list_):
        return ErrorManager._clone(list_[0]) if len(list_) > 0 else None

    @staticmethod
    def get(label=None, message=None, level=None):
        errors = ErrorManager.all()
        if not label and not message and not level:
            return ErrorManager.last()
        if label is not None:
            return ErrorManager._first_or_none([error for error in errors if error.label == str(label)])
        if message is not None:
            return [error for error in errors if error.message == str(message)]
        if level is not None:
            return [error for error in errors if error.level == level]

    @staticmethod
    def create_from_exception(e):
        return Error(e.__class__.__name__, str(e), logging.ERROR, params=None)

    @staticmethod
    def _clone(error):
        result = copy.copy(error)
        result.params = None
        return result

    def __init__(self):
        pass


class Error():
    def __init__(self, label, message, level, params=None):
        self.label = label
        self.message = message
        self.level = level
        self.params = params

    def __str__(self):
        result = self.message or self.label
        if isinstance(self.params, dict):
            result = self.message.format(**self.params)
        elif isinstance(self.params, list):
            result = self.message.format(*self.params)
        return result

    objects = ErrorManager


class ValidationException(Exception):
    pass


class NoMetaDataSheetException(ValidationException):
    pass


class WrongVisitException(ValidationException):
    pass


class WrongSiteException(ValidationException):
    pass


class MissingDataSheetException(ValidationException):
    pass


class MissingVegetationVisit(ValidationException):
    pass


class NoDataForModelException(ValidationException):
    pass


class TooManyDataForModelException(ValidationException):
    pass


class FieldErrorException(ValidationException):
    pass


class InvalidSpeciesNameException(ValidationException):
    pass


class DataSheetModelMapping():
    """
    This class keep tracks
    """

    def __init__(self, model,
                 sheet_name=None,
                 top_left_column=1,
                 top_left_row=1,
                 transpose=False,
                 mandatory=True,
                 unique=True,
                 formulas=None,
                 extra_validation_func=None):
        self.model = model
        default_sheet_name = self.model._meta.verbose_name_plural.capitalize() if self.model is not None else None
        self.sheet_name = sheet_name or default_sheet_name
        self.top_left_column = top_left_column
        self.top_left_row = top_left_row
        self.transpose = transpose
        self.mandatory = mandatory  # True = must have at least one row
        self.unique = unique  # True = must have at most one row
        self.formulas = formulas if formulas is not None else {}
        self.extra_validation_func = extra_validation_func

    def get_table_data(self, worksheet):
        return util_xls.TableData(worksheet,
                                  top_left_column=self.top_left_column,
                                  top_left_row=self.top_left_row,
                                  transpose=self.transpose)

    @property
    def next_col_direction(self):
        return 'down' if self.transpose else 'right'

    @property
    def next_row_direction(self):
        return 'right' if self.transpose else 'down'


DATASHEET_META_MAPPING = DataSheetModelMapping(None, sheet_name="Meta", top_left_row=2, top_left_column=2,
                                               transpose=True, mandatory=True,
                                               unique=True)
DATASHEET_MAIN_MODELS_MAPPING = [
    DataSheetModelMapping(SiteCharacteristic,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=True, mandatory=False,
                          unique=True),
]

DATASHEET_VEGETATION_MODELS_MAPPING = [
    DataSheetModelMapping(VegetationVisit,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=True, mandatory=True,
                          unique=True),
    DataSheetModelMapping(StratumSpecies,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False),
    DataSheetModelMapping(TransectObservation,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=True, mandatory=False,
                          unique=True,
                          extra_validation_func=ground_cover_validate),
    DataSheetModelMapping(TransectDistinctChanges,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False),
    DataSheetModelMapping(BasalBitterlichObservation,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=True, mandatory=False,
                          unique=True),
    DataSheetModelMapping(ErosionPeg,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False),
    DataSheetModelMapping(PegObservation,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False),
    DataSheetModelMapping(GroundCoverSummary,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False,
                          extra_validation_func=ground_cover_validate),
    DataSheetModelMapping(StratumSummary,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False),
    DataSheetModelMapping(DisturbanceIndicator,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=True, mandatory=False,
                          unique=True),
    DataSheetModelMapping(PlantObservation,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False),
    DataSheetModelMapping(BiodiversityIndicator,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=True, mandatory=False,
                          unique=True),
]

DATASHEET_ANIMALS_MODELS_MAPPING = [
    DataSheetModelMapping(Trap,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False,
                          extra_validation_func=lat_lon_validate),
    DataSheetModelMapping(AnimalObservation,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=True,
                          unique=False,
                          # formulas={'net_weight': {'formula': '=%s-%s', 'parameters': ('gross_weight', 'bag_weight')}}
                          ),
    DataSheetModelMapping(OpportunisticObservation,
                          sheet_name=None, top_left_row=1, top_left_column=1, transpose=False, mandatory=False,
                          unique=False,
                          extra_validation_func=lat_lon_validate),
]

DATASHEET_MODELS_MAPPING = \
    DATASHEET_MAIN_MODELS_MAPPING + \
    DATASHEET_VEGETATION_MODELS_MAPPING + \
    DATASHEET_ANIMALS_MODELS_MAPPING


def get_mapping_for_model(model):
    return next((m for m in DATASHEET_MODELS_MAPPING if m.model == model), None)


class SiteDataFileValidator():
    def __init__(self, sv_file):
        self.sv_file = sv_file

    def validate(self):
        """
        :return: a SiteVisitDataFileError if error or None if all is good
        """
        result = None
        errors = []
        # WARNING! It is important to NOT open the workbook in read_only mode, it breaks the utilities in utils_openpyxl
        wb = load_workbook(self.sv_file.file.path, read_only=False)
        try:
            site = parse_site_raise(wb, self.sv_file.visit)
            self.sv_file.site = site
            self.sv_file.save()
        except Exception as e:
            errors.append(Error.objects.create_from_exception(e))

        for model in DATASHEET_MODELS_MAPPING:
            validator = ModelValidator(model)
            errors += validator.validate(wb)
        if errors:
            result = SiteVisitDataFileError(file=self.sv_file)
            result.message = "\n".join([str(error) for error in errors])
            result.save()
        return result


class ModelValidator():
    def __init__(self, model_mapping):
        self.model_mapping = model_mapping
        self.model = model_mapping.model

    def validate(self, wb):
        errors = []
        # find worksheet
        ws_name = self.model_mapping.sheet_name
        ws = util_xls.get_sheet(wb, ws_name)
        if ws is None:
            error = Error.objects.get(label='missing_data_datasheet')
            error.params = {'sheet_name': ws_name}
            errors.append(error)
        else:
            # worksheet found. Validate data
            table_data = self.model_mapping.get_table_data(ws)
            if len(table_data.rows) == 0 and self.model_mapping.mandatory:
                error = ErrorManager.get(label="no_data_for_model")
                error.params = {'model': ws_name}
                errors.append(error)
            if len(table_data.rows) > 1 and self.model_mapping.unique:
                error = ErrorManager.get(label="too_many_data_for_model")
                error.params = {'model': ws_name}
                errors.append(error)
            else:
                row_index = 0

                for row_data in table_data.by_rows():
                    row_index += 1
                    errors += self.validate_row(row_data, row_index, ws)

                    if callable(self.model_mapping.extra_validation_func):
                        try:
                            self.model_mapping.extra_validation_func(row_data)
                        except Exception as e:
                            error = Error.objects.get(label="extra_validation_error")
                            params = {
                                'datasheet': ws.title,
                                'row': row_index + 1,  # add 1 to match the worksheet number
                                'msg': str(e)
                            }
                            error.params = params
                            errors.append(error)

        return errors

    def validate_row(self, row_data, row_index, ws):
        """
        :param row_data: expect an array of tuples [(col1, val1), (col2, val2),...]
        :return: list of errors
        """
        errors = []
        fields = util_model.get_datasheet_fields_for_model(self.model)
        for field in fields:
            field_sheet_name = util_model.get_datasheet_field_name(field)
            field_value = None
            for name, value in row_data:
                if name == field_sheet_name:
                    field_value = value
                    break
            try:
                to_field_value_raise(field, field_value, commit=False, row_data=row_data)
            except Exception as e:
                error = Error.objects.get(label="field_error")
                params = {
                    'datasheet': ws.title,
                    'col': field_sheet_name,
                    'row': row_index + 1,  # add 1 to match the worksheet number
                    'msg': str(e)
                }
                error.params = params
                errors.append(error)

        return errors


class SiteVisitDataBuilder():
    def __init__(self, sv_file):
        self.sv_file = sv_file
        # We expect the SiteVisitDataFile to have a visit and a site foreign key
        if self.sv_file.visit is None:
            raise Exception("Can't build data from a SiteVisitDataFile without a visit")
        if self.sv_file.site is None:
            raise Exception("Can't build data from a SiteVisitDataFile without a site")
        # WARNING! It is important to NOT open the workbook in read_only mode, it breaks the utilities in utils_openpyxl
        self.wb = load_workbook(self.sv_file.file.path, read_only=False)
        # this is the top object. Every object instance should have this has a parent.
        self.site_visit = self._create_site_data_obj()

    def build_all(self, keep_invalid=False):
        try:
            self._build_main_objects()._build_vegetation_objects()._build_animal_objects()
        except Exception as e:
            logger.exception(e)
            if not keep_invalid:
                # deleting the top level object site_visit should trigger a cascade
                # TODO: test the cascade delete
                self.site_visit.delete()
            raise e

        return self.site_visit

    def _build_main_objects(self):
        for mapping in DATASHEET_MAIN_MODELS_MAPPING:
            parser = ModelBuilder(mapping, site_visit=self.site_visit)
            parser.build(self.wb, commit=True)
        return self

    def _build_vegetation_objects(self):
        # top level of all vegetation tables are a vegetation visit
        parser = ModelBuilder(get_mapping_for_model(VegetationVisit), site_visit=self.site_visit)
        vegetation_visits = parser.build(self.wb, commit=True)
        vegetation_visit = vegetation_visits[0] if len(vegetation_visits) > 0 else None
        # parse all children
        children_model_mappings = [m for m in DATASHEET_VEGETATION_MODELS_MAPPING if m.model != VegetationVisit]
        for mapping in children_model_mappings:
            parser = ModelBuilder(mapping, site_visit=self.site_visit, vegetation_visit=vegetation_visit)
            parser.build(self.wb, commit=True)
        return self

    def _build_animal_objects(self):
        for mapping in DATASHEET_ANIMALS_MODELS_MAPPING:
            parser = ModelBuilder(mapping, site_visit=self.site_visit)
            parser.build(self.wb, commit=True)
        return self

    def _create_site_data_obj(self):
        site_visit_obj = SiteVisit.objects.create(
            visit=self.sv_file.visit,
            site=self.sv_file.site,
            data_file=self.sv_file
        )
        return site_visit_obj


class ModelBuilder():
    """
    Class in charge of creating the db object of a given model from a spreadsheet.
    """

    def __init__(self, model_mapping, site_visit, **kwargs):
        self.model_mapping = model_mapping
        self.model = model_mapping.model
        self.site_visit = site_visit
        self.initial_values = kwargs

    def build(self, wb, commit=False):
        objects = []
        # find worksheet
        ws_name = self.model_mapping.sheet_name
        ws = util_xls.get_sheet(wb, ws_name)
        if ws is None and self.model_mapping.mandatory:
            raise MissingDataSheetException
        else:
            # worksheet found. Parse rows
            table_data = self.model_mapping.get_table_data(ws)
            if len(table_data.rows) == 0 and self.model_mapping.mandatory:
                message = "No data found for {model}".format(model=self.model._meta.verbose_name)
                raise NoDataForModelException(message)
            if len(table_data.rows) > 1 and self.model_mapping.unique:
                message = "Too many rows for {model}. It should only contains one row (or column)" \
                    .format(model=self.model._meta.verbose_name)
                raise TooManyDataForModelException(message)
            else:
                for row_data in table_data.by_rows():
                    obj = self._create_obj(row_data)
                    if commit:
                        obj.save()
                    objects.append(obj)

        return objects

    def _create_obj(self, row_data):
        """
        :param row_data: [(field.verbose_name, value) ...]
        """
        obj = self._init_object()
        for field_vname, value in row_data:
            field = util_model.get_field_by_verbose_name(self.model, field_vname)
            if field is not None:
                field_value = to_field_value_raise(field, value, site_visit=self.site_visit, commit=True,
                                                   row_data=row_data)
                ModelBuilder._set_field(obj, field, field_value, save=False)
            else:
                # TODO: what to do with extra columns. Add it in a HStore field?
                logger.warning("Found a column {col} that doesn't match a field for model {model}"
                               .format(col=field_vname, model=self.model))
        return obj

    def _init_object(self):
        obj = self.model(**self.initial_values)
        if hasattr(obj, 'site_visit_id'):
            obj.site_visit = self.site_visit
        return obj

    @staticmethod
    def _set_field(obj, field, value, save=False):
        setattr(obj, field.name, value)
        if save:
            obj.save()


def parse_site_raise(wb, visit):
    """
    Parse the Meta worksheet and check for visit and site validity
    :return: The site object in db that matches what is in the Meta or throw an exception
    """
    meta = MetaData.parse_workbook(wb)
    # check that the visit match
    if meta.visit_name is None or meta.visit_name != visit.name:
        raise WrongVisitException("Visit ID not found or doesn't match.")
    if meta.site_code is None:
        raise WrongSiteException("No site code found!")
    sites = visit.sites.filter(site_code=meta.site_code)
    if len(sites) == 0:
        raise WrongSiteException("The site {site} is not a valid site for the visit {visit}"
                                 .format(site=meta.site_code, visit=visit.name))
    if len(sites) > 1:
        raise WrongSiteException("There is more than one site {site} for the visit {visit}"
                                 .format(site=meta.site_code, visit=visit.name))
    return sites.first()


def to_field_value_raise(field, value, commit=True, site_visit=None, row_data=None):
    if is_blank(value):
        if util_model.is_mandatory(field):
            message = "Mandatory field with no data {field}.".format(field=field.verbose_name)
            raise ValidationException(message)
    # Every next conversion functions should handle a blank/None value
    if util_model.is_lookup_field(field):
        return to_lookup_raise(field, value, commit=commit)
    if util_model.has_choices(field):
        return to_choice_raise(field, value)
    if util_model.is_boolean_field(field):
        return to_boolean_raise(value)
    if util_model.is_integer_field(field):
        return to_integer_raise(value)
    if util_model.is_float_field(field):
        return to_float_raise(value)
    if util_model.is_date_field(field):
        return to_date_raise(value)
    if util_model.is_string_field(field):
        return to_string(value)
    if util_model.is_species_observation_field(field):
        return to_species_observation_raise(value, site_visit, commit=commit, row_data=row_data)
    return value


def to_species_observation_raise(value, site_visit=None, commit=True, row_data=None):
    """
    Validate the supplied species name value. Rules:
    *
    Create a SpeciesObservation object with an input_name = value
    This function also try to get a name_id reference from the species data reference
    If a no name_id is returned it means that the species name is probably invalid but we still create the object
    without raising an exception
    :param value:
    :param site_visit:
    :param commit: save the object if True
    :return: the SpeciesObservation object
    """
    if is_blank(value):
        message = "Species is blank."
        raise FieldErrorException(message)
    value = str(value)
    manager = SpeciesObservationManager(value, row_data=row_data)

    return manager.create(site_visit=site_visit, commit=commit)


class SpeciesObservationManager:
    """
    For BIOSYS-117: add extra column validation status and uncertainty for every species field.
    """

    def __init__(self, species, row_data=None):
        self.model = SpeciesObservation
        self.species = species
        self.row_data = row_data

    @property
    def must_validate(self):
        return is_blank(self.validation_status)

    @property
    def validation_status(self):
        return self._get_field_value('validation_status')

    @property
    def uncertainty(self):
        return self._get_field_value('uncertainty')

    def validate_raise(self):
        # a valid name_id is returned if a matching species is found
        name_id = None
        if self.must_validate:
            # Remove multiple spaces, tabs, newlines, leading/trailing spaces.
            value = ' '.join(self.species.split()).capitalize()

            # Query the database for an extra match.
            try:
                sp = Species.objects.get(species_name=value)
                # found!
                name_id = sp.name_id
            except Species.MultipleObjectsReturned:  # Should never happen.
                raise FieldErrorException('Species matches multiple records.')
            except Species.DoesNotExist:
                sp = None
            if sp is None:
                # Split value on space into at most three parts: [genus, species, remainder]
                sp = value.split(' ', 2)
                # At this point, we have uncertainty about species (but still might match genus).
                if sp and len(sp) == 1:  # Only supplied genus.
                    raise FieldErrorException('No species supplied (input "sp." if unknown).')
                else:
                    # No match on exact species name; try "Genus sp."
                    genus = sp[0]
                    species = sp[1]
                    if species != 'sp.':
                        msg = 'Species mismatch: "{}" (input "sp." if species is unknown).'.format(value)
                        raise FieldErrorException(msg)
                    qs = Species.objects.similar_name(genus, similarity=0.55)  # NOTE: may require adjustment.
                    qs = list(qs)
                    if len(qs) == 0:  # No match for genus.
                        msg = 'Unknown genus: {} (input "sp." if species is unknown)'.format(genus)
                        raise FieldErrorException(msg)
                    else:  # At least one matching genus was found.
                        # At this point we allow the value; should be: Genus sp. <OPTIONAL EXTRAS>
                        pass
        return name_id

    def create(self, site_visit=None, commit=True):
        name_id = self.validate_raise()
        input_name = self.species
        obj = SpeciesObservation(input_name=input_name)
        if name_id:
            obj.name_id = name_id
        if self.uncertainty:
            obj.uncertainty = self.uncertainty
        if self.validation_status:
            obj.validation_status = self.validation_status
        if site_visit:
            obj.site_visit = site_visit
        if commit:
            obj.save()
        return obj

    def _get_field_value(self, field_name):
        field = self.model._meta.get_field(field_name)
        return self._get_value(util_model.get_datasheet_field_name(field))

    def _get_value(self, key):
        result = None
        if self.row_data:
            for k, v in self.row_data:
                if k == key:
                    return v
        return result


def to_species_name_id(species_name):
    """
    This where we hook to the species reference db to grab a species name_id
    from a species name
    :param species_name:
    :return:
    """
    obj = Species.objects.filter(species_name=species_name).first()
    return obj.name_id if obj is not None else None


def to_string(value):
    """
    By convention None is converted to empty string
    :param value:
    :return:
    """
    return str(value) if value is not None else ""


def to_date_raise(value, default=None):
    if is_blank(value):
        return default
    if isinstance(value, datetime.datetime):
        value = value.date()
    value = str(value).strip()
    valid_formats = settings.DATE_INPUT_FORMATS
    date = None
    for format_ in valid_formats:
        try:
            date = datetime.datetime.strptime(value, format_)
            break
        except ValueError:
            pass
    if date is None:
        message = "{value} is not a valid date.".format(value=value)
        raise ValidationException(message)
    else:
        return date


def to_float_raise(value, default=None):
    try:
        return float(str(value).strip()) if not is_blank(value) else default
    except ValueError:
        raise ValidationException("{value} is not a valid float".format(value=value))


def to_integer_raise(value, default=None):
    try:
        return int(str(value).strip()) if not is_blank(value) else default
    except ValueError:
        raise ValidationException("{value} is not a valid integer".format(value=value))


def to_boolean_raise(value, default=False):
    if is_blank(value):
        return default
    true_values = ['true', 'yes', 'x', 'y']
    false_values = ['false', 'no', 'n', '']
    value = str(value).strip().lower()
    if value in true_values:
        return True
    if value in false_values:
        return False
    raise ValidationException("{value} is not a valid boolean value. Should be one of {accepted_values}"
                              .format(value=value, accepted_values=true_values + false_values))


def to_choice_raise(field, value):
    """
    Rules:
        validate only against the display_name, the second part of a Django choice (internal, display_name)
        case insensitive
    """
    if not util_model.has_choices(field):
        raise FieldErrorException("The field {field} has no choices".format(field.verbose_name))
    if is_blank(value):
        return field.default
    value = str(value)
    choices = util_model.get_field_choices(field)
    return to_model_choice(choices, value)


def to_model_choice(choices, value):
    """
    Rules:
        validate only against the display_name, the second part of a Django choice (internal, display_name)
        case insensitive
        :param choices: a model choice as an array of tuples
        :param value:
    """
    choice = next((c[0] for c in choices if c[1].lower() == value.lower()), None)
    if choice is None:
        message = "{value} not an authorized choice. Should be one of: {values}" \
            .format(value=value, values=[str(c[1]) for c in choices])
        raise FieldErrorException(message)
    return choice


def to_lookup_raise(field, value, commit=True, default=None):
    """
    Rules:
        validate only for 'strict' lookups
        validate against code or values (case insensitive)
        deprecated lookups are rejected.
    """
    if is_blank(value):
        return default
    if not util_model.is_lookup_field(field):
        raise FieldErrorException("{field} is not a lookup".format(field=field))
    lookup_model = field.related_model
    value = str(value) if value is not None else None  # operate on string only
    # search for lookup code first and value after (case insensitive)
    lookup = lookup_model.objects.filter(code__iexact=value).first() \
             or lookup_model.objects.filter(value__iexact=value).first()
    if lookup is None:
        # can't find. If lookup is strict, it's an error
        if util_model.is_strict_lookup_field(field):
            codes = [str(c) for c in util_model.get_field_lookup_codes(field)]
            values = [str(v) for v in util_model.get_field_lookup_values(field)]
            accepted_values = codes + values
            message = "{value} is not an authorized lookup value for {field}. Should be one of: {values}" \
                .format(value=value, field=field.verbose_name, values=accepted_values)
            raise FieldErrorException(message)
        elif value is not None and len(value.strip()) > 0:
            # if not strict we add the new value (capitalized) in the lookup table.
            lookup = lookup_model(value=value.title())
            if commit:
                lookup.save()
    elif lookup.deprecated:
        message = "{value} is a deprecated value for {field}" \
            .format(value=value, field=field.verbose_name)
        raise FieldErrorException(message)

    return lookup


class MetaData():
    @staticmethod
    def parse_file(sv_file):
        path = sv_file.file.path if isinstance(sv_file, SiteVisitDataFile) else sv_file
        wb = load_workbook(path, read_only=False)
        return MetaData.parse_workbook(wb)

    @staticmethod
    def parse_workbook(wb):
        mapping = DATASHEET_META_MAPPING
        meta_ws = util_xls.get_sheet(wb, mapping.sheet_name)
        if meta_ws is None:
            raise NoMetaDataSheetException("Meta sheet not found.")
        key = util_model.get_field_verbose_name(Visit, 'name')
        reading_direction = 'right' if mapping.transpose else 'down'
        visit_name = util_xls.get_value_for_key(meta_ws, key, direction=reading_direction)
        key = util_model.get_field_verbose_name(Site, 'site_code')
        site_code = util_xls.get_value_for_key(meta_ws, key, direction=reading_direction)
        return MetaData(visit_name, site_code)

    def __init__(self, visit_name, site_code):
        # visit_name and site_code must be string or None
        self.visit_name = str(visit_name).strip() if not is_blank(visit_name) else None
        self.site_code = str(site_code).strip() if not is_blank(site_code) else None
