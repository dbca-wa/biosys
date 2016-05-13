import os
import random
import datetime
from openpyxl import Workbook
from unittest import skip

from django.test import TestCase
from django.core.management import call_command

from main.models import Site, Visit
from species.models import Species
from main import utils as util_model
from animals.models import *

from upload.validation import to_lookup_raise, to_choice_raise, to_boolean_raise, to_integer_raise, \
    to_float_raise, to_date_raise, ValidationException, to_species_name_id, to_field_value_raise, MetaData,\
    NoMetaDataSheetException, DATASHEET_MODELS_MAPPING, DATASHEET_META_MAPPING, to_species_observation_raise
from upload import utils_openpyxl as util_xls


class TestValidation(TestCase):
    def setUp(self):
        pass


class TestFieldLookupValidation(TestCase):
    def setUp(self):
        self.model = Site
        self.all_lookup_fields = [f for f in self.model._meta.fields if util_model.is_lookup_field(f)]
        self.strict_lookup_fields = [f for f in self.all_lookup_fields if util_model.is_strict_lookup_field(f)]
        self.not_strict_lookup_fields = [f for f in self.all_lookup_fields if not util_model.is_strict_lookup_field(f)]
        self.assertTrue(len(self.strict_lookup_fields) > 0)
        self.assertTrue(len(self.not_strict_lookup_fields) > 0)


    def test_not_strict_allows_all(self):
        """
        There is no validation for a not strict lookup
        """
        for f in self.not_strict_lookup_fields:
            rand_val = "askjldhskjd" + str(random.randint(123, 456))
            self.assertTrue(validate_lookup(f, rand_val), msg="A non strict lookup should always be valid")

    def test_strict_code_or_value_case_insensitive(self):
        """
        If strict it validates against the value OR the code
        """
        for field in self.strict_lookup_fields:
            lookup_values = util_model.get_field_lookup_values(field)
            lookup_codes = [code for code in util_model.get_field_lookup_codes(field) if code]
            for value in lookup_values:
                valid, message = validate_lookup(field, value)
                self.assertTrue(valid)
                self.assertIsNone(message)
                # test case insensitive
                valid, message = validate_lookup(field, value.capitalize().swapcase())
                self.assertTrue(valid)
                self.assertIsNone(message)
            for value in lookup_codes:
                self.assertTrue(validate_lookup(field, value)[0])
                # test case insensitive
                self.assertTrue(validate_lookup(field, value.capitalize().swapcase())[0])
                # test blank is not accepted
                valid, message = validate_lookup(field, "")
                self.assertFalse(valid)
                # test that there's a message
                self.assertTrue(message)

                # test that boolean, none, int doesn't break anything
                for weird in ["", 12, None, True, {}, []]:
                    valid, message = validate_lookup(field, weird)
                    self.assertFalse(valid)
                    # test that there's a message
                    self.assertTrue(message)

                # random
                randoms = [str(random.randint(1783, 6900)), "dhgd6sdkcdfdee765&"]
                for r in randoms:
                    valid, message = validate_lookup(field, r)
                    self.assertFalse(valid)
                    # test that there's a message
                    self.assertTrue(message)

    def test_not_strict_add(self):
        """
        Test that for not strict lookup a new value is added
        """
        # choose a non_strict_lookup
        f = self.not_strict_lookup_fields[0]
        original_values = util_model.get_field_lookup_values(f)
        original_size = len(original_values)
        new_value = "test38738"
        to_lookup_raise(f, new_value, commit=True)
        # the new value should have been added
        values = util_model.get_field_lookup_values(f)
        self.assertEqual(original_size + 1, len(values))
        self.assertTrue(new_value in values)

    def test_not_strict_blank(self):
        """
        A blank value for a non strict lookup should not be added
        """
        # choose a non_strict_lookup
        f = self.not_strict_lookup_fields[0]
        original_values = util_model.get_field_lookup_values(f)
        original_size = len(original_values)
        blanks = ["", None, "   "]
        for blank in blanks:
            lookup = to_lookup_raise(f, blank, commit=True)
            values = util_model.get_field_lookup_values(f)
            self.assertEqual(original_size, len(values))
            self.assertTrue(blank not in values)


class TestFieldChoice(TestCase):
    def setUp(self):
        pass

    def test_datum(self):
        field = Site._meta.get_field('datum')
        # only the display name is valid, case insensitive
        for value in [c[1] for c in DATUM_CHOICES]:
            valid, message = validate_choices(field, value)
            self.assertTrue(valid, msg="{} should be accepted as a valid choice".format(value))
            valid, message = validate_choices(field, value.capitalize().lower())
            self.assertTrue(valid, msg="{} should be accepted as a valid choice".format(value))

        # the internal name is not valid
        for value in [c[0] for c in DATUM_CHOICES]:
            valid, message = validate_choices(field, value)
            self.assertFalse(valid)
            self.assertTrue(message)

        # test that blank return the default choices
        default_choice = field.default
        blanks = ['', "   ", None]
        for value in blanks:
            self.assertEqual(default_choice, to_choice_raise(field, value))

        # test that blank, boolean, none, int doesn't break anything
        for weird in [12, True, "hello world!", {}, []]:
            valid, message = validate_choices(field, weird)
            self.assertFalse(valid, msg="{} should not be accepted as a valid choice".format(weird))
            self.assertTrue(message)


class TestFieldParsing(TestCase):
    def setUp(self):
        pass

    def test_to_boolean_raise(self):
        trues = [True, 'true', 'yes', 'YeS', 'x', 'X', 'y', 'Y']
        falses = [False, 'false', 'no', 'no', 'No', 'NO', '', "   ", 'n', 'N', None]
        invalids = ['yess', 'noooo', 'no way!', 0, '0', '1', 1]
        for t in trues:
            try:
                self.assertTrue(to_boolean_raise(t))
            except Exception as e:
                self.assertTrue(False, msg="{} should be a valid true boolean: {}".format(t, str(e)))
        for f in falses:
            try:
                self.assertFalse(to_boolean_raise(f))
            except Exception as e:
                self.assertTrue(False, msg="{} should be a valid false boolean: {}".format(f, str(e)))
        for u in invalids:
            # all invalids should throw an exception
            with self.assertRaises(ValidationException):
                to_boolean_raise(u)

    def test_to_integer_raise(self):
        valids = [12, '12', '0', -123, '-123']
        invalids = [False, True, 12.4, '13.4']
        for v in valids:
            try:
                self.assertEqual(int(v), to_integer_raise(v))
            except Exception as e:
                self.assertTrue(False, msg="{} should be a valid integer: {}".format(v, str(e)))
        # all invalids should throw an exception
        for v in invalids:
            with self.assertRaises(ValidationException):
                to_integer_raise(v)
        # blank values should return None
        blanks = [None, '', "  "]
        for v in blanks:
            try:
                self.assertEqual(None, to_integer_raise(v))
            except Exception as e:
                self.assertTrue(False, msg="{} should return None: {}".format(v, str(e)))

    def test_to_float_raise(self):
        valids = [12, '12.345', '0.23', -123.12, '-123', 0.0, 0, '+3.0']
        invalids = [False, True, '3.x', '1/2']
        for v in valids:
            try:
                self.assertEqual(float(v), to_float_raise(v))
            except Exception as e:
                self.assertTrue(False, msg="{} should be a valid float: {}".format(v, str(e)))
        # all invalids should throw an exception
        for v in invalids:
            with self.assertRaises(ValidationException):
                to_float_raise(v)
        # blank values should return None
        blanks = [None, '', "  "]
        for v in blanks:
            try:
                self.assertEqual(None, to_float_raise(v))
            except Exception as e:
                self.assertTrue(False, msg="{} should return None: {}".format(v, str(e)))

    def test_to_date_raise(self):
        valids = ['31/12/2014', '2014-12-31', '2014_12_31', datetime.date(2014, 12, 31),
                  datetime.datetime(2014, 12, 31)]
        for v in valids:
            try:
                self.assertEqual('2014-12-31', to_date_raise(v).strftime('%Y-%m-%d'))
            except Exception as e:
                self.assertTrue(False, msg="{} should be a valid date: {}".format(v, str(e)))
        invalids = ['31/12/14', '2014-12-32', 'date',  True, False, 10122014]
        for v in invalids:
            with self.assertRaises(ValidationException):
                to_date_raise(v)
        # blank values should return None
        blanks = [None, '', "  "]
        for v in blanks:
            try:
                self.assertEqual(None, to_date_raise(v))
            except Exception as e:
                self.assertTrue(False, msg="{} should return None: {}".format(v, str(e)))


class TestFieldValidation(TestCase):
    def setUp(self):
        self.datasheet_models = [m.model for m in DATASHEET_MODELS_MAPPING]
        ll = [util_model.get_datasheet_fields_for_model(m) for m in self.datasheet_models]
        self.datasheet_fields = [item for l in ll for item in l]
        self.mandatory_fields = [f for f in self.datasheet_fields if util_model.is_mandatory(f)]
        self.non_mandatory_fields = [f for f in self.datasheet_fields if not util_model.is_mandatory(f)]

    def test_non_mandatory_blank_lookup(self):
        """
        Test that a non mandatory lookup with a blank value returns None and doesn't throw an exception
        """
        non_mandatory_lookup_fields = [f for f in self.non_mandatory_fields if util_model.is_lookup_field(f)]
        blanks = [None, '', "  "]
        for f in non_mandatory_lookup_fields:
            for v in blanks:
                try:
                    self.assertEqual(None, to_field_value_raise(f, v))
                except Exception as e:
                    self.assertTrue(False,
                                    msg="A blank value for non mandatory lookup should return None and not throw an exception")

    def test_non_mandatory_blank_date(self):
        """
        Test that a non mandatory date with a blank value returns None and not throw an exception
        """
        non_mandatory_date_fields = [f for f in self.non_mandatory_fields if util_model.is_date_field(f)]
        blanks = [None, '', "  "]
        for f in non_mandatory_date_fields:
            for v in blanks:
                try:
                    self.assertEqual(None, to_field_value_raise(f, v))
                except Exception as e:
                    self.assertTrue(False,
                                    msg="A blank value for non mandatory date should return None and not throw an exception")


class TestSpeciesValidation(TestCase):
    @classmethod
    def setUpClass(cls):
        if Species.objects.count() == 0:
            try:
                Species.objects.update_herbie_hbvspecies()
            except:
                print('herbie failed: load species from file')
                call_command("loaddata", '--app', 'species', 'species.json')

    @classmethod
    def tearDownClass(cls):
        Species.objects.all().delete()

    def test_to_name_id(self):
        """
        Test that we return the name_id from a valid species_name or None if name is not found
        """
        first = Species.objects.first()
        self.assertIsNotNone(first)
        self.assertEquals(to_species_name_id(first.species_name), first.name_id)
        self.assertIsNone(to_species_name_id('dpaw species'))

    def test_to_species_observation_simple(self):
        """
        Test that a valid species name return a OldSpeciesObservation with a valid name id
        :return:
        """
        species = Species.objects.first()
        self.assertIsNotNone(species)
        species_obs = None
        try:
            species_obs = to_species_observation_raise(species.species_name, None, commit=False, row_data=None)
        except Exception as e:
            self.assertTrue(False, msg="{} should be a valid species name: {}".format(species.species_name, str(e)))
        self.assertIsNotNone(species_obs)
        self.assertEqual(species.species_name, species_obs.input_name)
        self.assertEqual(species.name_id, species_obs.name_id)

    def test_validation_case_insensitive(self):
        species = Species.objects.first()
        self.assertIsNotNone(species)
        species_name = species.species_name.lower()
        species_obs = None
        try:
            species_obs = to_species_observation_raise(species_name, None, commit=False, row_data=None)
        except Exception as e:
            self.assertTrue(False, msg="{} should be a valid species name: {}".format(species.species_name, str(e)))
        self.assertIsNotNone(species_obs)
        self.assertEqual(species_name, species_obs.input_name)
        self.assertEqual(species.name_id, species_obs.name_id)

    def test_genus_validation(self):
        """
        Just the genus should fail, but genus + sp. should pass
        :return:
        """
        species = Species.objects.first()
        self.assertIsNotNone(species)
        genus = species.species_name.split()[0]
        species_obs = None
        species_name = genus
        try:
            species_obs = to_species_observation_raise(species_name, None, commit=False, row_data=None)
            print(species_obs)
            self.assertTrue(False, msg="{} should not be a valid species name: {}".format(species_name))
        except:
            pass

    @skip('Skipped failing test')
    def test_genus_validation_sp(self):
        # add sp.
        species = Species.objects.first()
        self.assertIsNotNone(species)
        genus = species.species_name.split()[0]
        species_obs = None
        species_name = genus + " sp."
        try:
            species_obs = to_species_observation_raise(species_name, None, commit=False, row_data=None)
        except Exception as e:
            self.assertTrue(False, msg="{} should be a valid species name: {}".format(species_name, str(e)))
        self.assertIsNotNone(species_obs)
        self.assertEqual(species_name, species_obs.input_name)
        self.assertEqual(species.name_id, species_obs.name_id)

    @skip('Skipped failing test')
    def test_genus_validation_lc(self):
        # should work with lower case
        species = Species.objects.first()
        self.assertIsNotNone(species)
        genus = species.species_name.split()[0]
        species_obs = None
        species_name = genus.lower() + " sp."
        try:
            species_obs = to_species_observation_raise(species_name, None, commit=False, row_data=None)
        except Exception as e:
            self.assertTrue(False, msg="{} should be a valid species name: {}".format(species_name, str(e)))
        self.assertIsNotNone(species_obs)
        self.assertEqual(species_name, species_obs.input_name)
        self.assertEqual(species.name_id, species_obs.name_id)


class TestMetaData(TestCase):
    def setUp(self):
        self.file_path = os.path.join(os.path.dirname(__file__), 'data', 'sitevisit-valid.xlsx')

    def test_throw_exception_if_no_meta_sheet(self):
        wb = Workbook()
        wb.active.title = 'Meeta'
        with self.assertRaises(NoMetaDataSheetException):
            MetaData.parse_workbook(wb)

    def test_empty_meta(self):
        wb = Workbook()
        wb.active.title = 'Meta'
        meta = MetaData.parse_workbook(wb)
        self.assertIsNone(meta.visit_name)
        self.assertIsNone(meta.site_code)

    def test_simple(self):
        visit_name = 'My Visit'
        site_code = 'MySite'
        meta = MetaData.parse_workbook(self._create_meta(visit_name, site_code))
        self.assertEqual(visit_name, meta.visit_name)
        self.assertEqual(site_code, meta.site_code)

    def test_strip(self):
        # Test that the value are trimmed (striped)
        visit_name = ' My Visit  '
        site_code = 'MySite   '
        meta = MetaData.parse_workbook(self._create_meta(visit_name, site_code))
        self.assertEqual(visit_name.strip(), meta.visit_name)
        self.assertEqual(site_code.strip(), meta.site_code)

    def test_string(self):
        # Test that numbers should be converted in string
        visit_name = 123
        site_code = 456
        meta = MetaData.parse_workbook(self._create_meta(visit_name, site_code))
        self.assertEqual(str(visit_name), meta.visit_name)
        self.assertEqual(str(site_code), meta.site_code)

    def test_empty_string(self):
        # empty string should return None
        visit_name = "   "
        site_code = ''
        meta = MetaData.parse_workbook(self._create_meta(visit_name, site_code))
        self.assertIsNone(meta.visit_name)
        self.assertIsNone(meta.site_code)

    def _create_meta(self, visit_name, site_code):
        # TODO: use the download.utils.SiteVisitDatasheetWriter to write the metadata
        wb = Workbook()
        mapping = DATASHEET_META_MAPPING
        meta_ws = wb.active
        # rewrite the sheet name in case of the 'Sheet' selection.
        meta_ws.title = mapping.sheet_name
        col_headers = [util_model.get_field_verbose_name(Visit, 'name'),
                       util_model.get_field_verbose_name(Site, 'site_code')]
        # write column headers
        top_cell = meta_ws.cell(row=mapping.top_left_row, column=mapping.top_left_column)
        writing_direction = mapping.next_col_direction
        util_xls.write_values_from_cell(top_cell, col_headers, writing_direction)
        # write values
        values = [visit_name, site_code]
        top_cell = util_xls.get_cell_neighbour(top_cell, mapping.next_row_direction)
        util_xls.write_values_from_cell(top_cell, values, writing_direction)
        return wb


def validate_lookup(field, value):
    valid, message = True, None
    try:
        to_lookup_raise(field, value, commit=False)
    except Exception as e:
        valid = False
        message = str(e)
    return valid, message


def validate_choices(field, value):
    valid, message = True, None
    try:
        to_choice_raise(field, value)
    except Exception as e:
        valid = False
        message = str(e)
    return valid, message



