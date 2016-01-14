import random

from django.test import TestCase
from openpyxl import Workbook

import upload.utils_openpyxl as utils


class TestUtilsOpenPyXL(TestCase):
    def setUp(self):
        pass

    def test_get_sheet(self):
        """
        Test that accessing a sheet that doesn't exists return None and not throw an exception
        """
        wb = Workbook()
        wb.create_sheet(0, 'Test')
        self.assertTrue('Test' in utils.get_sheet_titles(wb))
        ws = utils.get_sheet(wb, 'Test')
        self.assertIsNotNone(ws)
        wb.remove_sheet(ws)
        self.assertIsNone(utils.get_sheet(wb, 'Test'))

    def test_get_sheet_case_insensitive(self):
        """
        By default it is case insensitive
        """
        wb = Workbook()
        wb.create_sheet(0, 'Test')
        self.assertTrue('Test' in utils.get_sheet_titles(wb))
        ws = utils.get_sheet(wb, 'test')
        self.assertIsNotNone(ws)
        self.assertIsNone(utils.get_sheet(wb, 'test', case_insensitive=False))
        wb.remove_sheet(ws)
        self.assertIsNone(utils.get_sheet(wb, 'test'))


class TestTableData(TestCase):
    def setUp(self):
        pass

    def test_simple(self):
        """
        Just create a table 10*4 and read it from the top left of the sheet
        """
        wb = Workbook()
        ws = wb.active
        col_headers = range(1, 11)  # ten columns 1...10
        ws.append(col_headers)
        rows = []
        for i in range(0, 4):
            row = [random.randint(0, 100) for r in col_headers]
            ws.append(row)
            rows.append(row)

        tb = utils.TableData(ws)  # default top-left = (0,0)
        self.assertEqual(len(col_headers), len(tb.column_headers))
        self.assertEqual(col_headers, tb.column_headers)
        self.assertEqual(len(rows), len(tb.rows))
        self.assertEqual(rows, tb.rows)
        # test the by_columns
        by_columns = tb.by_columns()
        self.assertEqual(len(col_headers), len(by_columns))
        for col, values in by_columns:
            self.assertTrue(col in col_headers)
            self.assertEqual(len(rows), len(values))
        # test the by_rows
        by_rows = tb.by_rows()
        self.assertEqual(len(rows), len(by_rows))
        for row in by_rows:
            self.assertTrue(len(row), len(col_headers))

    def test_transpose(self):
        wb = Workbook()
        ws = wb.active
        # generate a transposed table at (10,10)
        start_col = 10
        start_row = 10
        col_headers = ['Col1', 'Col2', 'Col3', 'Col4', 'Col5']
        rows = [
            ['r11', 'r12', 'r13', 'r14', 'r15'],
            ['r21', 'r22', 'r23', 'r24', 'r25'],
            ['r31', 'r32', 'r33', 'r34', 'r35'],
            ['r41', 'r42', 'r43', 'r44', 'r45'],
            ['r51', 'r52', 'r53', 'r54', 'r55'],
            ['r61', 'r62', 'r63', 'r64', 'r65']
        ]
        col_index = start_col
        row_index = start_row
        # col headers
        for h in col_headers:
            ws.cell(row=row_index, column=col_index).value = h
            row_index += 1
        # rows (they are columns)
        col_index = start_col + 1
        row_index = start_row
        for row in rows:
            for cell in row:
                ws.cell(row=row_index, column=col_index).value = cell
                row_index += 1
            col_index += 1
            row_index = start_row

        tb = utils.TableData(ws,
                             top_left_row=start_col, top_left_column=start_row,
                             transpose=True)
        self.assertEqual(len(col_headers), len(tb.column_headers))
        self.assertEqual(col_headers, tb.column_headers)
        self.assertEqual(len(rows), len(tb.rows))
        self.assertEqual(rows, tb.rows)
        # test the by_columns
        by_columns = tb.by_columns()
        self.assertEqual(len(col_headers), len(by_columns))
        for col, values in by_columns:
            self.assertTrue(col in col_headers)
            self.assertEqual(len(rows), len(values))
        # test the by_rows
        by_rows = tb.by_rows()
        self.assertEqual(len(rows), len(by_rows))
        for row in by_rows:
            self.assertTrue(len(row), len(col_headers))
