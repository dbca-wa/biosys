from copy import copy
from django.core.urlresolvers import reverse
from django.http import HttpResponse, HttpResponseBadRequest
from django.views.generic import TemplateView, View
from openpyxl import Workbook
import os
import requests
from urllib import urlencode

from main.admin import user_can_approve
from main.models import SiteCharacteristic
from main.views import datasheet_schema
from vegetation.models import (
    StratumSpecies, TransectObservation, TransectDistinctChanges,
    BasalBitterlichObservation, ErosionPeg, PegObservation, GroundCoverSummary,
    StratumSummary, DisturbanceIndicator, PlantObservation, BiodiversityIndicator
    )
from animals.models import Trap, AnimalObservation, OpportunisticObservation


class ReportView(TemplateView):
    """Template view to allow filtering and analysis of data.
    """
    template_name = 'publish/report.html'

    def get_context_data(self, **kwargs):
        context = super(ReportView, self).get_context_data(**kwargs)
        context['publish_report_view'] = True
        # Extra context for the JavaScript in the template.
        if user_can_approve(self.request.user):
            context['is_approver'] = 'true'
        else:
            context['is_approver'] = 'false'
        return context


class DownloadView(View):
    """A basic view to return a spreadsheet of data related to SiteVisit objects.
    """
    def get(self, request):
        # Get the datasheet schema rules.
        schema = datasheet_schema()
        # Generate a blank Excel workbook.
        wb = Workbook()
        wb.remove_sheet(wb.get_active_sheet())  # Remove the initial sheet.
        # Get any query parameters to filter the data.
        query_params = dict(request.GET.iteritems())
        # If present, use the request session for querying the API endpoints.
        if 'sessionid' in request.COOKIES:
            cookies = {'sessionid': request.COOKIES['sessionid']}
        else:
            cookies = {}

        for model in [
                SiteCharacteristic, StratumSpecies, TransectObservation,
                TransectDistinctChanges, BasalBitterlichObservation, ErosionPeg,
                PegObservation, GroundCoverSummary, StratumSummary,
                DisturbanceIndicator, PlantObservation, BiodiversityIndicator,
                Trap, AnimalObservation, OpportunisticObservation]:
            # First, find the matching object rule in the datasheet schema.
            model_schema = next(i for i in schema['models'] if i['class_name'] == model._meta.object_name)

            # From the model schema, build our list of table headers.
            headers = ['Project', 'Site', 'Visit', 'Status']
            for i in model_schema['fields']:
                headers.append(i['datasheet_name'])

            # Create a new blank worksheet for the model data.
            ws = wb.create_sheet(title=model._meta.verbose_name_plural.capitalize())

            # Write the column headers to the new worksheet.
            for col, value in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = value

            # Prepend site_visit or vegetation_visit__site_visit to query params, as req'd.
            params_copy = copy(query_params)
            for k in params_copy.iterkeys():
                if k.startswith('site'):  # Need to alter the key
                    if model_schema['module'].split('.')[0] == 'vegetation':
                        params_copy['vegetation_visit__site_visit__' + k] = params_copy.pop(k)
                    else:
                        params_copy['site_visit__' + k] = params_copy.pop(k)

            # Build an absolute URL to the model API endpoint.
            url = reverse(
                'api_dispatch_list',
                kwargs={
                    'resource_name': model._meta.object_name.lower(),
                    'api_name': 'v1'}
            )
            url = url + '?' + urlencode(params_copy)
            url = 'http://localhost:{}'.format(os.environ['PORT']) + url

            # For each model, query the API endpoint for data.
            resp = requests.get(url, cookies=cookies)
            if not resp.status_code == 200:
                return HttpResponseBadRequest('Download failed')
            objects = resp.json()['objects']

            # For each object, write attributes to the worksheet in the order
            # defined by the model schema.
            for row, obj in enumerate(objects, 2):  # Start at row 2
                obj_data = []
                # Project, site and visit depends on the model class.
                if model_schema['module'].split('.')[0] == 'vegetation':
                    obj_data.append(obj['vegetation_visit']['site_visit']['site']['project']['title'])
                    obj_data.append(obj['vegetation_visit']['site_visit']['site']['site_code'])
                    obj_data.append(obj['vegetation_visit']['site_visit']['visit']['name'])
                    obj_data.append(obj['vegetation_visit']['site_visit']['data_status'])
                else:  # Other app models just follow through SiteVisit.
                    obj_data.append(obj['site_visit']['site']['project']['title'])
                    obj_data.append(obj['site_visit']['site']['site_code'])
                    obj_data.append(obj['site_visit']['visit']['name'])
                    obj_data.append(obj['site_visit']['data_status'])

                # Iterate over the object fields in order of the model schema.
                for field in model_schema['fields']:
                    if field['is_lookup']:  # Handle lookup values
                        # a lookup can be null
                        value = obj[field['name']]['value'] if obj[field['name']] is not None else ""
                        obj_data.append(value)
                    elif field['is_species_name']:  # Handle species names
                        # species is normally never null
                        value = obj[field['name']]['input_name'] if obj[field['name']] is not None else ""
                        obj_data.append(value)
                    elif field['is_extra_species_attribute']:
                        if obj['species']:
                            value = obj['species'][field['name']]
                            obj_data.append(value)
                    elif field['is_boolean']:  #Handle boolean values
                        if obj[field['name']] is not None:
                            value = "Yes" if obj[field['name']] else "No"
                        else:
                            value = ""
                        obj_data.append(value)
                    else:
                        obj_data.append(obj[field['name']])

                # Finally, write all object values to a new worksheet row.
                for col, val in enumerate(obj_data, 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = val

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=biosys.xlsx'
        wb.save(response)  # Save the workbook contents to the response.
        return response
