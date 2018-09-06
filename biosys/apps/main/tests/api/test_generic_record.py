import json
import re
from os import path

from django.core.urlresolvers import reverse
from django.utils import six
from openpyxl import load_workbook
from openpyxl.cell import Cell
from rest_framework import status

from main.models import Dataset, Record
from main.tests import factories
from main.tests.api import helpers
from main.tests.test_data_package import clone


class TestPermissions(helpers.BaseUserTestCase):
    """
    Test Permissions
    Get: authenticated
    Update: admin, custodians
    Create: admin, custodians
    Delete: admin, custodians
    """
    def setUp(self):
        super(TestPermissions, self).setUp()
        self.ds_1_rows = [
            ['What', 'When', 'Who'],
            ['Something', '2018-02-01', 'me']
        ]
        self.ds_1 = self._create_dataset_and_records_from_rows(self.ds_1_rows)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = Record.objects.filter(dataset=self.ds_1).first()
        self.record_1.site = factories.SiteFactory.create(project=self.project_1)
        self.record_1.save()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))

        self.ds_2_rows = [
            ['Who', 'Height', 'Weight', 'Comments'],
            ['Me', '1.86', '80', 'I wish']
        ]
        self.ds_2 = self._create_dataset_and_records_from_rows(self.ds_1_rows)
        self.ds_2.project = self.project_2
        self.ds_2.save()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

    def test_get(self):
        urls = [
            reverse('api:record-list'),
            reverse('api:record-detail', kwargs={'pk': self.record_1.pk})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.get(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.get(url).status_code,
                    status.HTTP_200_OK
                )

    def test_create(self):
        """
        Admin and custodians
        :return:
        """
        urls = [reverse('api:record-list')]
        rec = self.record_1
        data = {
            "dataset": rec.dataset.pk,
            "site": rec.site.pk,
            "data": rec.data
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                count = Record.objects.count()
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )
                self.assertEqual(Record.objects.count(), count + 1)

    def test_bulk_create(self):
        """
        Cannot create bulk with this end point
        :return:
        """
        urls = [reverse('api:record-list')]
        rec = self.record_1
        data = [
            {
                "dataset": rec.dataset.pk,
                "site": rec.site.pk,
                "data": rec.data
            },
            {
                "dataset": rec.dataset.pk,
                "site": rec.site.pk,
                "data": rec.data
            }
        ]
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client,
                          self.admin_client, self.custodian_1_client],
            "allowed": []
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_400_BAD_REQUEST, status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                count = Record.objects.count()
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )
                self.assertEqual(Record.objects.count(), count + len(data))

    def test_update(self):
        """
        admin + custodian of project for site 1
        :return:
        """
        rec = self.record_1
        previous_data = clone(rec.data)
        updated_data = clone(previous_data)
        updated_data['Longitude'] = '118.78'
        urls = [reverse('api:record-detail', kwargs={'pk': rec.pk})]
        data = {
            "data": updated_data,
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.patch(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                rec.data = previous_data
                rec.save()
                self.assertEqual(
                    client.patch(url, data, format='json').status_code,
                    status.HTTP_200_OK
                )
                rec.refresh_from_db()
                self.assertEqual(rec.data, updated_data)

    def test_delete(self):
        """
        Currently admin + custodian
        :return:
        """
        rec = self.record_1
        urls = [reverse('api:record-detail', kwargs={'pk': rec.pk})]
        data = None
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.delete(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                rec.save()
                count = Dataset.objects.count()
                self.assertEqual(
                    client.delete(url, data, format='json').status_code,
                    status.HTTP_204_NO_CONTENT
                )
                self.assertTrue(Dataset.objects.count(), count - 1)

    def test_options(self):
        urls = [
            reverse('api:record-list'),
            reverse('api:record-detail', kwargs={'pk': 1})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.options(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        # authenticated
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.options(url).status_code,
                    status.HTTP_200_OK
                )


class TestDataValidation(helpers.BaseUserTestCase):

    def setUp(self):
        super(TestDataValidation, self).setUp()
        self.ds_1_rows = [
            ['What', 'When', 'Who'],
            ['Something', '2018-02-01', 'me']
        ]
        self.ds_1 = self._create_dataset_and_records_from_rows(self.ds_1_rows)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = Record.objects.filter(dataset=self.ds_1).first()
        self.record_1.site = factories.SiteFactory.create(project=self.project_1)
        self.record_1.save()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))

        self.ds_2_rows = [
            ['Who', 'Height', 'Weight', 'Comments'],
            ['Me', '1.86', '80', 'I wish']
        ]
        self.ds_2 = self._create_dataset_and_records_from_rows(self.ds_1_rows)
        self.ds_2.project = self.project_2
        self.ds_2.save()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

    def test_create_one_happy_path(self):
        """
        Test the create of one record
        :return:
        """
        # grab one existing an re-inject it
        record = self.record_1
        data = {
            "dataset": record.dataset.pk,
            "data": record.data
        }
        url = reverse('api:record-list')
        client = self.data_engineer_1_client
        count = Record.objects.count()
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(Record.objects.count(), count + 1)

    def test_empty_not_allowed(self):
        record = self.record_1
        data = {
            "dataset": record.dataset.pk,
            "data": {}
        }
        url = reverse('api:record-list')
        # set strict mode
        url = helpers.set_strict_mode(url)
        client = self.custodian_1_client
        count = Record.objects.count()
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(Record.objects.count(), count)

    def test_create_column_not_in_schema(self):
        """
        Test that if we introduce a column not in the the dataset it will not validate
        :return:
        """
        record = self.record_1
        incorrect_data = clone(record.data)
        incorrect_data['Extra Column'] = "Extra Value"
        data = {
            "dataset": record.dataset.pk,
            "data": incorrect_data
        }
        url = reverse('api:record-list')
        # set strict mode
        url = helpers.set_strict_mode(url)
        client = self.custodian_1_client
        count = Record.objects.count()
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(Record.objects.count(), count)

    def test_update_column_not_in_schema(self):
        """
        Test that if we introduce a column not in the the dataset it will not validate
        :return:
        """
        record = self.record_1
        incorrect_data = clone(record.data)
        incorrect_data['Extra Column'] = "Extra Value"
        data = {
            "dataset": record.dataset.pk,
            "data": incorrect_data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        # set strict mode
        url = helpers.set_strict_mode(url)
        client = self.custodian_1_client
        count = Record.objects.count()
        self.assertEqual(
            client.put(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(Record.objects.count(), count)
        self.assertEqual(
            client.patch(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(Record.objects.count(), count)


class TestSiteExtraction(helpers.BaseUserTestCase):

    def setUp(self):
        super(TestSiteExtraction, self).setUp()
        self.site_1 = factories.SiteFactory(project=self.project_1, code='COT')
        self.ds_1_rows = [
            ['What', 'When', 'Who', 'Site'],
        ]
        self.ds_1 = self._create_dataset_from_rows(self.ds_1_rows)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        # add a site code foreign key
        self.schema_1 = self.ds_1.schema_data
        helpers.add_model_field_foreign_key_to_schema(
            self.schema_1,
            {
                'schema_field': 'Site',
                'model': 'Site',
                'model_field': 'code'
            }
        )
        self.ds_1.data_package = helpers.create_data_package_from_schema(self.schema_1)
        self.ds_1.save()
        self.ds_1.refresh_from_db()
        self.assertTrue(self.ds_1.schema.has_fk_for_model('Site'))
        # create one record with site
        self.record_1 = self._create_record(
            self.custodian_1_client,
            self.ds_1,
            {
                'What': 'Something',
                'When': '2018-06-30',
                'Site': self.site_1.code
            })
        self.assertEqual(self.record_1.site.pk, self.site_1.pk)

    def test_create_with_site(self):
        """
        The descriptor contains a foreign key to the site.
        Test that the site is extracted from the data
        :return:
        """
        # clear all records
        Record.objects.all().delete()
        self.assertEqual(Record.objects.count(), 0)
        record = self.record_1
        data = {
            "dataset": record.dataset.pk,
            "data": record.data
        }
        schema = self.ds_1.schema
        self.assertTrue(schema.has_fk_for_model('Site'))
        expected_site = self.site_1
        url = reverse('api:record-list')
        client = self.custodian_1_client
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(Record.objects.count(), 1)
        self.assertEqual(Record.objects.first().site, expected_site)

    def test_update_site(self):
        record = self.record_1
        site = self.site_1
        self.assertEqual(self.record_1.site.pk, self.site_1.pk)
        # need to test if the site belongs to the dataset project or the update won't happen
        self.assertIsNotNone(site)
        self.assertTrue(site.project == record.dataset.project)
        # update site value
        schema = record.dataset.schema
        site_column = schema.get_fk_for_model('Site').data_field
        r_data = record.data
        r_data[site_column] = site.code
        data = {
            "data": r_data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        self.assertEqual(
            client.patch(url, data, format='json').status_code,
            status.HTTP_200_OK
        )
        record.refresh_from_db()
        self.assertEqual(record.site, site)


class TestExport(helpers.BaseUserTestCase):

    def setUp(self):
        super(TestExport, self).setUp()
        self.ds_1_rows = [
            ['What', 'When', 'Who'],
            ['Something', '2018-02-01', 'me']
        ]
        self.ds_1 = self._create_dataset_and_records_from_rows(self.ds_1_rows)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = Record.objects.filter(dataset=self.ds_1).first()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))

        self.ds_2_rows = [
            ['Who', 'Height', 'Weight', 'Comments'],
            ['Me', '1.86', '80', 'I wish']
        ]
        self.ds_2 = self._create_dataset_and_records_from_rows(self.ds_1_rows)
        self.ds_2.project = self.project_2
        self.ds_2.save()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

    def test_happy_path_no_filter(self):
        client = self.custodian_1_client
        dataset = self.ds_1
        all_records = Record.objects.filter(dataset=dataset)
        self.assertTrue(all_records.count() > 0)
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        # check headers
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=DatasetName_YYYY_MM_DD-HHMMSS.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEqual(ext, '.xlsx')
        filename.startswith(dataset.name)
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named from dataset
        sheet_names = wb.sheetnames
        self.assertEqual(1, len(sheet_names))
        self.assertEqual(dataset.name, sheet_names[0])
        ws = wb[dataset.name]
        rows = list(ws.rows)
        expected_records = Record.objects.filter(dataset=dataset)
        self.assertEqual(len(rows), expected_records.count() + 1)
        headers = [c.value for c in rows[0]]
        schema = dataset.schema
        # all the columns of the schema should be in the excel
        self.assertEqual(schema.headers, headers)

    def test_permission_ok_for_not_custodian(self):
        """Export is a read action. Should be authorised for every logged-in user."""
        client = self.custodian_2_client
        dataset = self.ds_1
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_permission_denied_if_not_logged_in(self):
        """Must be logged-in."""
        client = self.anonymous_client
        dataset = self.ds_1
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)

    def test_excel_type(self):
        schema_fields = [
            {
                "name": "string",
                "type": "string",
                "format": "default"
            },
            {
                "name": "number",
                "type": "number",
                "format": "default"
            },
            {
                "name": "integer",
                "type": "integer",
                "format": "default"
            },
            {
                "name": "date",
                "type": "date",
                "format": "any"
            },
            {
                "name": "datetime",
                "type": "datetime",
                "format": "any"
            },
            {
                "name": "boolean",
                "type": "boolean",
                "format": "default"
            }
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        project = self.project_1
        client = self.data_engineer_1_client
        dataset = self._create_dataset_with_schema(project, client, schema)

        # create one record
        record_data = {
            "string": "string",
            "number": 12.3,
            "integer": 456,
            "date": "21/06/2017",
            "datetime": "13/04/2017 15:55",
            "boolean": 'yes'
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        # load workbook
        wb = load_workbook(six.BytesIO(resp.content))
        ws = wb[dataset.name]
        rows = list(ws.rows)
        self.assertEqual(len(rows), 2)
        cells = rows[1]
        string, number, integer, date, datetime, boolean = cells
        # excel type are string, number or boolean
        self.assertEqual(string.data_type, Cell.TYPE_STRING)
        self.assertEqual(number.data_type, Cell.TYPE_NUMERIC)
        self.assertEqual(integer.data_type, Cell.TYPE_NUMERIC)
        self.assertEqual(date.data_type, Cell.TYPE_NUMERIC)
        self.assertEqual(datetime.data_type, Cell.TYPE_NUMERIC)
        self.assertEqual(boolean.data_type, Cell.TYPE_BOOL)


class TestSchemaValidation(helpers.BaseUserTestCase):

    def assert_create_dataset(self, schema):
        try:
            return self._create_dataset_with_schema(
                self.project_1,
                self.data_engineer_1_client,
                schema,
                dataset_type=Dataset.TYPE_GENERIC
            )
        except Exception as e:
            self.fail('Species Observation dataset creation failed for schema {schema}'.format(
                schema=schema
            ))

    def test_not_required_date_with_format_any(self):
        """
        field of type date, not required with format any should not raise an error when received a empty string
        see https://decbugs.com/view.php?id=6928
        """
        schema_fields = [
            {
                "name": "DateAny",
                "type": "date",
                "format": "any",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "DateTimeAny",
                "type": "datetime",
                "format": "any",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "DateDefault",
                "type": "date",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "DateTimeDefault",
                "type": "datetime",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            }
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        dataset = self.assert_create_dataset(schema)
        records = [
            ['DateAny', 'DateTimeAny', 'DateDefault', 'DateTimeDefault'],
            [None, None, None, None],
            ['', '', '', ''],
            ['  ', '   ', '  ', '  '],
        ]
        resp = self._upload_records_from_rows(records, dataset_pk=dataset.pk, strict=True)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_required_date_with_format_any(self):
        """
        field of type date, required with format should
        """
        schema_fields = [
            {
                "name": "DateAny",
                "type": "date",
                "format": "any",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            }
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        dataset = self.assert_create_dataset(schema)
        records = [
            ['DateAny', 'What'],
            [None, 'something'],
            ['', 'something'],
            ['   ', 'something'],
        ]
        resp = self._upload_records_from_rows(records, dataset_pk=dataset.pk, strict=True)
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        received = resp.json()
        self.assertIsInstance(received, list)
        self.assertEqual(len(received), 3)
        # this what an report should look like
        expected_row_report = {
            'row': 3,
            'errors': {'DateAny': 'Field "DateAny" has constraint "required" which is not satisfied for value "None"'},
            'warnings': {}}
        for row_report in received:
            self.assertIn('errors', row_report)
            errors = row_report.get('errors')
            self.assertIn('DateAny', errors)
            msg = errors.get('DateAny')
            self.assertEqual(msg, expected_row_report['errors']['DateAny'])


class TestPatch(helpers.BaseUserTestCase):

    def test_patch_validated(self):
        """
        Test that we can patch just the 'validated' flag
        :return:
        """
        rows = [
            ['What', 'Comments'],
            ['Chubby bat', 'It is huge!']
        ]
        dataset = self._create_dataset_and_records_from_rows(rows)
        self.assertEqual(dataset.type, Dataset.TYPE_GENERIC)
        records = dataset.record_set.all()
        record = records.last()
        self.assertIsNotNone(record)
        self.assertFalse(record.validated)
        previous_data = json.dumps(record.data)
        # patch
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        payload = {
            'validated': True
        }
        resp = client.patch(url, payload)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertTrue(record.validated)
        self.assertTrue(json.dumps(record.data), previous_data)

    def test_patch_locked(self):
        """
        Test that we can patch just the 'locked' flag
        :return:
        """
        rows = [
            ['What', 'Comments'],
            ['Chubby bat', 'It is huge!']
        ]
        dataset = self._create_dataset_and_records_from_rows(rows)
        self.assertEqual(dataset.type, Dataset.TYPE_GENERIC)
        records = dataset.record_set.all()
        record = records.last()
        self.assertIsNotNone(record)
        self.assertFalse(record.locked)
        previous_data = json.dumps(record.data)
        # patch
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        payload = {
            'locked': True
        }
        resp = client.patch(url, payload)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertTrue(record.locked)
        self.assertTrue(json.dumps(record.data), previous_data)


class TestForeignKey(helpers.BaseUserTestCase):
    """
    Tests are about a dataset schema declaring one of its field as a foreign key to another dataset schema field
    When declaring a FK in the schema you have to define a resource name. This name is supposed to be a dataset name.
    or the resource name.
    Note: in order to get children records the parent dataset schema MUST declare a 'primaryKey' property.
    """

    def setUp(self):
        super(TestForeignKey, self).setUp()
        # delete all datasets
        Dataset.objects.all().delete()

    def test_fk_with_dataset_name(self):
        """
        Test that the parent children works when the declared FK refer to a dataset name.
        :return:
        """
        # Create a parent dataset with some records
        parent_dataset = self._create_dataset_and_records_from_rows([
            ['Survey ID', 'Where', 'When', 'Who'],
            ['ID-001', 'King\'s Park', '2018-07-15', 'Tim Reynolds'],
            ['ID-002', 'Cottesloe', '2018-07-11', 'SLB'],
            ['ID-003', 'Somewhere', '2018-07-13', 'Phil Bill']
        ])
        parent_dataset.data_package['resources'][0]['schema']['primaryKey'] = 'Survey ID'
        parent_dataset.save()
        parent_records = parent_dataset.record_set.all()
        self.assertEqual(parent_records.count(), 3)

        # Create a child/related schema
        child_schema = helpers.create_schema_from_fields([
            {
                "name": "Survey ID",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "Comments",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
            }
        ])
        # declaring a foreign key pointing to the dataset name
        foreign_keys = [{
            'fields': 'Survey ID',
            'reference': {
                'fields': 'Survey ID',
                'resource': parent_dataset.name
            }
        }]
        child_schema['foreignKeys'] = foreign_keys
        child_dataset = self._create_dataset_with_schema(
            self.project_1,
            self.data_engineer_1_client,
            child_schema
        )
        self.assertIsNotNone(child_dataset)
        # post some records for survey ID--001
        rows = [
            ['Survey ID', 'What', 'Comments'],
            ['ID-001', 'Canis lupus', 'doggy, doggy'],
            ['ID-001', 'A frog', 'kiss'],
            ['ID-001', 'A tooth brush', 'I should stop drinking'],
        ]
        self._upload_records_from_rows(rows, child_dataset.id, strict=False)
        children_records = child_dataset.record_set.all()
        self.assertEqual(children_records.count(), 3)

        # test serialisation of parent records
        id_001 = parent_records.filter(data__contains={'Survey ID': 'ID-001'}).first()
        expected_children_ids = [r.id for r in children_records]
        expected_parent_id = None
        self.assertIsNotNone(id_001)
        url = reverse('api:record-detail', kwargs={'pk': id_001.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        id_002 = parent_records.filter(data__contains={'Survey ID': 'ID-002'}).first()
        expected_children_ids = []
        expected_parent_id = None
        self.assertIsNotNone(id_002)
        url = reverse('api:record-detail', kwargs={'pk': id_002.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        id_003 = parent_records.filter(data__contains={'Survey ID': 'ID-003'}).first()
        expected_children_ids = []
        expected_parent_id = None
        self.assertIsNotNone(id_003)
        url = reverse('api:record-detail', kwargs={'pk': id_003.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        #
        # test serialisation of children records
        # they all have the same parent and no children
        #
        expected_children_ids = None
        expected_parent_id = id_001.pk
        client = self.custodian_1_client
        for record in children_records:
            url = reverse('api:record-detail', kwargs={'pk': record.pk})
            resp = client.get(url)
            self.assertEqual(resp.status_code, status.HTTP_200_OK)
            data = resp.json()
            self.assertEqual(data['children'], expected_children_ids)
            self.assertEqual(data['parent'], expected_parent_id)

    def test_fk_with_dataset_code(self):
        """
        This is the same test as above but this time the foreign key is declared with the parent dataset code instead
        of name
        :return:
        """
        # Create a parent dataset with some records
        parent_dataset = self._create_dataset_and_records_from_rows([
            ['Survey ID', 'Where', 'When', 'Who'],
            ['ID-001', 'King\'s Park', '2018-07-15', 'Tim Reynolds'],
            ['ID-002', 'Cottesloe', '2018-07-11', 'SLB'],
            ['ID-003', 'Somewhere', '2018-07-13', 'Phil Bill']
        ])
        parent_dataset.data_package['resources'][0]['schema']['primaryKey'] = 'Survey ID'
        parent_dataset.save()
        parent_dataset.code = 'Survey'
        parent_dataset.save()
        parent_records = parent_dataset.record_set.all()
        self.assertEqual(parent_records.count(), 3)

        # Create a child/related schema
        child_schema = helpers.create_schema_from_fields([
            {
                "name": "Survey ID",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "Comments",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
            }
        ])
        # declaring a foreign key pointing to the dataset name
        foreign_keys = [{
            'fields': 'Survey ID',
            'reference': {
                'fields': 'Survey ID',
                'resource': parent_dataset.code
            }
        }]
        child_schema['foreignKeys'] = foreign_keys
        child_dataset = self._create_dataset_with_schema(
            self.project_1,
            self.data_engineer_1_client,
            child_schema
        )
        self.assertIsNotNone(child_dataset)
        # post some records for survey ID--001
        rows = [
            ['Survey ID', 'What', 'Comments'],
            ['ID-001', 'Canis lupus', 'doggy, doggy'],
            ['ID-001', 'A frog', 'kiss'],
            ['ID-001', 'A tooth brush', 'I should stop drinking'],
        ]
        self._upload_records_from_rows(rows, child_dataset.id, strict=False)
        children_records = child_dataset.record_set.all()
        self.assertEqual(children_records.count(), 3)

        # test serialisation of parent records
        id_001 = parent_records.filter(data__contains={'Survey ID': 'ID-001'}).first()
        expected_children_ids = [r.id for r in children_records]
        expected_parent_id = None
        self.assertIsNotNone(id_001)
        url = reverse('api:record-detail', kwargs={'pk': id_001.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        id_002 = parent_records.filter(data__contains={'Survey ID': 'ID-002'}).first()
        expected_children_ids = []
        expected_parent_id = None
        self.assertIsNotNone(id_002)
        url = reverse('api:record-detail', kwargs={'pk': id_002.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        id_003 = parent_records.filter(data__contains={'Survey ID': 'ID-003'}).first()
        expected_children_ids = []
        expected_parent_id = None
        self.assertIsNotNone(id_003)
        url = reverse('api:record-detail', kwargs={'pk': id_003.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        #
        # test serialisation of children records
        # they all have the same parent and no children
        #
        expected_children_ids = None
        expected_parent_id = id_001.pk
        client = self.custodian_1_client
        for record in children_records:
            url = reverse('api:record-detail', kwargs={'pk': record.pk})
            resp = client.get(url)
            self.assertEqual(resp.status_code, status.HTTP_200_OK)
            data = resp.json()
            self.assertEqual(data['children'], expected_children_ids)
            self.assertEqual(data['parent'], expected_parent_id)

    def test_fk_with_dataset_resource_name(self):
        """
        This is the same test as above but this time the foreign key is declared with the parent dataset resource name
        :return:
        """
        # Create a parent dataset with some records
        parent_dataset = self._create_dataset_and_records_from_rows([
            ['Survey ID', 'Where', 'When', 'Who'],
            ['ID-001', 'King\'s Park', '2018-07-15', 'Tim Reynolds'],
            ['ID-002', 'Cottesloe', '2018-07-11', 'SLB'],
            ['ID-003', 'Somewhere', '2018-07-13', 'Phil Bill']
        ])
        parent_dataset.data_package['resources'][0]['schema']['primaryKey'] = 'Survey ID'
        parent_dataset.save()
        parent_dataset.name = 'Survey'
        parent_dataset.code = 'SURV'
        parent_dataset.save()
        self.assertTrue(parent_dataset.resource_name)  # not None or empty string
        self.assertNotEqual(parent_dataset.resource_name, parent_dataset.name)
        self.assertNotEqual(parent_dataset.resource_name, parent_dataset.code)
        parent_records = parent_dataset.record_set.all()
        self.assertEqual(parent_records.count(), 3)

        # Create a child/related schema
        child_schema = helpers.create_schema_from_fields([
            {
                "name": "Survey ID",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "Comments",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
            }
        ])
        # declaring a foreign key pointing to the dataset name
        foreign_keys = [{
            'fields': 'Survey ID',
            'reference': {
                'fields': 'Survey ID',
                'resource': parent_dataset.resource_name
            }
        }]
        child_schema['foreignKeys'] = foreign_keys
        child_dataset = self._create_dataset_with_schema(
            self.project_1,
            self.data_engineer_1_client,
            child_schema
        )
        self.assertIsNotNone(child_dataset)
        # post some records for survey ID--001
        rows = [
            ['Survey ID', 'What', 'Comments'],
            ['ID-001', 'Canis lupus', 'doggy, doggy'],
            ['ID-001', 'A frog', 'kiss'],
            ['ID-001', 'A tooth brush', 'I should stop drinking'],
        ]
        self._upload_records_from_rows(rows, child_dataset.id, strict=False)
        children_records = child_dataset.record_set.all()
        self.assertEqual(children_records.count(), 3)

        # test serialisation of parent records
        id_001 = parent_records.filter(data__contains={'Survey ID': 'ID-001'}).first()
        expected_children_ids = [r.id for r in children_records]
        expected_parent_id = None
        self.assertIsNotNone(id_001)
        url = reverse('api:record-detail', kwargs={'pk': id_001.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        id_002 = parent_records.filter(data__contains={'Survey ID': 'ID-002'}).first()
        expected_children_ids = []
        expected_parent_id = None
        self.assertIsNotNone(id_002)
        url = reverse('api:record-detail', kwargs={'pk': id_002.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        id_003 = parent_records.filter(data__contains={'Survey ID': 'ID-003'}).first()
        expected_children_ids = []
        expected_parent_id = None
        self.assertIsNotNone(id_003)
        url = reverse('api:record-detail', kwargs={'pk': id_003.pk})
        client = self.custodian_1_client
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertEqual(sorted(data['children']), sorted(expected_children_ids))
        self.assertEqual(data['parent'], expected_parent_id)

        #
        # test serialisation of children records
        # they all have the same parent and no children
        #
        expected_children_ids = None
        expected_parent_id = id_001.pk
        client = self.custodian_1_client
        for record in children_records:
            url = reverse('api:record-detail', kwargs={'pk': record.pk})
            resp = client.get(url)
            self.assertEqual(resp.status_code, status.HTTP_200_OK)
            data = resp.json()
            self.assertEqual(data['children'], expected_children_ids)
            self.assertEqual(data['parent'], expected_parent_id)
