import re
from os import path

from django.test import override_settings
from openpyxl import load_workbook

from django.shortcuts import reverse
from django.utils import six
from rest_framework import status

from main.tests.api import helpers
# TODO: remove when python3
if six.PY2:
    import unicodecsv as csv
else:
    import csv


class TestFieldSelection(helpers.BaseUserTestCase):

    def _more_setup(self):
        # create some data with date and geometry
        self.rows = [
            ['What', 'When', 'Latitude', 'Longitude'],
            ['a big bird', '20018-01-24', -32.0, 115.75],
            ['a chubby bat ', '20017-12-24', -33.6, 116.678],
        ]
        self.dataset = self._create_dataset_and_records_from_rows(self.rows)

    def test_only_geometry(self):
        """
        Scenario: a web map user needs only the geometry field.
            Given some records with geometry are created
            And I request a get 'dataset-record' with fields=geometry
            Then it should return only the geometry field
        """
        # records are created in setup
        client = self.custodian_1_client
        url = reverse('api:dataset-records', kwargs={'pk': self.dataset.pk})
        query_params = {
            'fields': 'geometry'
        }
        resp = client.get(url, data=query_params, format='json')
        self.assertEqual(status.HTTP_200_OK, resp.status_code)
        records = resp.json()
        self.assertIsInstance(records, list)
        expected_record_count = len(self.rows) - 1
        self.assertEqual(len(records), expected_record_count)
        expected_fields = ['geometry']
        for record in records:
            self.assertIsInstance(record, dict)
            self.assertEqual(sorted(list(record.keys())), sorted(expected_fields))

    def test_geometry_and_id(self):
        """
        Scenario: a web map user needs only the geometry field and the record id to display an edit link.
            Given some records with geometry are created
            And I request a get 'dataset-record' with fields geometry and id
            Then it should return only the geometry and the id field
        """
        # records are created in setup
        client = self.custodian_1_client
        url = reverse('api:dataset-records', kwargs={'pk': self.dataset.pk})
        query_params = {
            'fields': ['geometry', 'id']
        }
        resp = client.get(url, data=query_params, format='json')
        self.assertEqual(status.HTTP_200_OK, resp.status_code)
        records = resp.json()
        self.assertIsInstance(records, list)
        expected_record_count = len(self.rows) - 1
        self.assertEqual(len(records), expected_record_count)
        expected_fields = ['geometry', 'id']
        for record in records:
            self.assertIsInstance(record, dict)
            self.assertEqual(sorted(list(record.keys())), sorted(expected_fields))

    def test_geometry_and_id_record_end_point(self):
        """
        Same as above but we hit the GET /records instead of GET/dataset/{pk}/records
        Scenario: a web map user needs only the geometry field and the record id to display an edit link.
            Given some records with geometry are created
            And I request a get 'record' with fields geometry and id
            Then it should return only the geometry and the id field
        """
        # records are created in setup
        client = self.custodian_1_client
        url = reverse('api:record-list')
        query_params = {
            'fields': ['geometry', 'id']
        }
        resp = client.get(url, data=query_params, format='json')
        self.assertEqual(status.HTTP_200_OK, resp.status_code)
        records = resp.json()
        self.assertIsInstance(records, list)
        expected_record_count = len(self.rows) - 1
        self.assertEqual(len(records), expected_record_count)
        expected_fields = ['geometry', 'id']
        for record in records:
            self.assertIsInstance(record, dict)
            # only key = geometry
            self.assertEqual(sorted(list(record.keys())), sorted(expected_fields))
            # request record individually
            url = reverse('api:record-detail', kwargs={'pk': record.get('id')})
            resp = client.get(url, data=query_params, format='json')
            self.assertEqual(status.HTTP_200_OK, resp.status_code)
            self.assertEqual(sorted(list(resp.json().keys())), sorted(expected_fields))

    def test_not_existing_field(self):
        """
        Scenario: asking for field that doesn't exists should not return an error but empty records
            Given some records with geometry are created
            And I request a get 'record' with a field 'field_with_typo'
            Then it should be successful
            And return records with no field
        """
        # records are created in setup
        client = self.custodian_1_client
        url = reverse('api:record-list')
        query_params = {
            'fields': ['field_with_typo']
        }
        resp = client.get(url, data=query_params, format='json')
        self.assertEqual(status.HTTP_200_OK, resp.status_code)
        records = resp.json()
        expected_fields = []
        for record in records:
            self.assertIsInstance(record, dict)
            self.assertEqual(sorted(list(record.keys())), sorted(expected_fields))

    def test_one_not_existing_field(self):
        """
        Scenario: asking for field that exists and one that doesn't exists should not return an error but the valid field
            Given some records with geometry are created
            And I request a get 'record' with a field 'geometry' and a field 'field_with_typo'
            Then it should be successful
            And return records with the geometry field
        """
        # records are created in setup
        client = self.custodian_1_client
        url = reverse('api:record-list')
        query_params = {
            'fields': ['geometry', 'field_with_typo']
        }
        resp = client.get(url, data=query_params, format='json')
        self.assertEqual(status.HTTP_200_OK, resp.status_code)
        records = resp.json()
        expected_fields = ['geometry']
        for record in records:
            self.assertIsInstance(record, dict)
            self.assertEqual(sorted(list(record.keys())), sorted(expected_fields))


class TestExcelFormat(helpers.BaseUserTestCase):

    @override_settings(EXPORTER_CLASS='main.api.exporters.DefaultExporter')
    def test_happy_path(self):
        expected_rows = [
            ['What', 'When', 'Latitude', 'Longitude'],
            ['a big bird in Cottesloe', '20018-01-24', -32.0, 115.75],
            ['a chubby bat somewhere', '20017-12-24', -33.6, 116.678],
            ['something in the null island', '2018-05-25', 0, 0]
        ]
        dataset = self._create_dataset_and_records_from_rows(expected_rows)
        client = self.custodian_1_client
        # ask for all records
        output = 'xlsx'
        url = reverse('api:record-list')
        query_params = {
            'dataset__id': dataset.pk,
            'output': output
        }
        resp = client.get(url, query_params)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=something.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEqual(ext, '.xlsx')
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named after the dataset
        expected_sheet_name = dataset.name
        sheet_names = wb.sheetnames
        self.assertEqual(1, len(sheet_names))
        self.assertEqual(sheet_names[0], expected_sheet_name)

        # check rows values
        ws = wb[expected_sheet_name]
        rows = list(ws.rows)
        # compare rows
        self.assertEqual(len(rows), len(expected_rows))
        for (expected_values, xlsx_row) in zip(expected_rows, rows):
            actual_values = [c.value for c in xlsx_row]
            self.assertEqual(expected_values, actual_values)


class TestCSVFormat(helpers.BaseUserTestCase):

    @override_settings(EXPORTER_CLASS='main.api.exporters.DefaultExporter')
    def test_happy_path(self):
        expected_rows = [
            ['What', 'When', 'Latitude', 'Longitude'],
            ['a big bird in Cottesloe', '20018-01-24', -32, 115.75],  # Note: if you put 32.0 the return will be '-32'
            ['a chubby bat somewhere', '20017-12-24', -33.6, 116.678],
            ['something in the null island', '2018-05-25', 0, 0]
        ]
        dataset = self._create_dataset_and_records_from_rows(expected_rows)
        client = self.custodian_1_client
        # ask for all records
        output = 'csv'
        url = reverse('api:record-list')
        query_params = {
            'dataset__id': dataset.pk,
            'output': output
        }
        resp = client.get(url, query_params)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.get('content-type'),
                         'text/csv')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=something.csv
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEqual(ext, '.csv')
        # read content
        reader = csv.reader(six.StringIO(resp.content.decode('utf-8')), dialect='excel')
        for expected_row, actual_row in zip(expected_rows, reader):
            expected_row_string = [str(v) for v in expected_row]
            self.assertEqual(actual_row, expected_row_string)


