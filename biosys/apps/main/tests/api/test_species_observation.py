import datetime
import re
from os import path
import json

from django.contrib.gis.geos import Point
from django.urls import reverse
from django.utils import timezone, six
from openpyxl import load_workbook
from rest_framework import status

from main.models import Dataset, Record
from main.tests.api import helpers
from main.tests.test_data_package import clone
from main.utils_species import NoSpeciesFacade


class TestPermissions(helpers.BaseUserTestCase):
    """
    Test Permissions
    Get: authenticated
    Update: admin, custodians
    Create: admin, custodians
    Delete: admin, custodians
    """
    species_facade_class = NoSpeciesFacade

    @staticmethod
    def schema_with_species_name():
        schema_fields = [
            {
                "name": "Species Name",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class
        project = self.project_1
        client = self.data_engineer_1_client
        schema = self.schema_with_species_name()
        self.ds_1 = self._create_dataset_with_schema(project, client, schema,
                                                     dataset_type=Dataset.TYPE_SPECIES_OBSERVATION)
        self.record_1 = self._create_default_record()

    def _create_default_record(self):
        ds = self.ds_1
        client = self.custodian_1_client
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        ds.record_queryset.delete()
        self.assertEqual(
            client.post(url, data=payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        return ds.record_queryset.first()

    def test_get(self):
        urls = [
            reverse('api:record-list'),
            reverse('api:record-detail', kwargs={'pk': self.record_1.pk})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.get(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.get(url).status_code,
                    status.HTTP_200_OK
                )

    def test_create(self):
        """
        Admin and custodians
        :return:
        """
        urls = [reverse('api:record-list')]
        ds = self.ds_1
        rec = self.record_1
        data = {
            "dataset": rec.dataset.pk,
            "data": rec.data,
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                count = ds.record_queryset.count()
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )
                self.assertEqual(ds.record_queryset.count(), count + 1)

    def test_bulk_create(self):
        """
        Cannot create bulk with this end point
        :return:
        """
        urls = [reverse('api:record-list')]
        rec = self.record_1
        ds = self.ds_1
        data = [
            {
                "dataset": rec.dataset.pk,
                "data": rec.data
            },
            {
                "dataset": rec.dataset.pk,
                "data": rec.data
            }
        ]
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client,
                          self.admin_client, self.custodian_1_client],
            "allowed": []
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_400_BAD_REQUEST, status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                count = ds.record_queryset.count()
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )
                self.assertEqual(ds.record_queryset.count(), count + len(data))

    def test_update(self):
        """
        admin + custodian of project for site 1
        :return:
        """
        rec = self.record_1
        previous_data = clone(rec.data)
        updated_data = clone(previous_data)
        updated_data['Longitude'] = '118.78'
        urls = [reverse('api:record-detail', kwargs={'pk': rec.pk})]
        data = {
            "data": updated_data,
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.patch(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                rec.data = previous_data
                rec.save()
                self.assertEqual(
                    client.patch(url, data, format='json').status_code,
                    status.HTTP_200_OK
                )
                rec.refresh_from_db()
                self.assertEqual(rec.data, updated_data)

    def test_delete(self):
        """
        Currently admin + custodian
        :return:
        """
        rec = self.record_1
        urls = [reverse('api:record-detail', kwargs={'pk': rec.pk})]
        data = None
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.delete(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                rec.save()
                count = Dataset.objects.count()
                self.assertEqual(
                    client.delete(url, data, format='json').status_code,
                    status.HTTP_204_NO_CONTENT
                )
                self.assertTrue(Dataset.objects.count(), count - 1)

    def test_options(self):
        urls = [
            reverse('api:record-list'),
            reverse('api:record-detail', kwargs={'pk': 1})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.options(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        # authenticated
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.options(url).status_code,
                    status.HTTP_200_OK
                )


class TestDataValidation(helpers.BaseUserTestCase):
    species_facade_class = NoSpeciesFacade

    @staticmethod
    def schema_with_species_name():
        schema_fields = [
            {
                "name": "Species Name",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        self.ds_1 = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema,
            dataset_type=Dataset.TYPE_SPECIES_OBSERVATION)

    def _create_default_record(self):
        ds = self.ds_1
        client = self.custodian_1_client
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        ds.record_queryset.delete()
        self.assertEqual(
            client.post(url, data=payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        return ds.record_queryset.first()

    def test_create_one_happy_path(self):
        """
        Test the create of one record
        :return:
        """
        ds = self.ds_1
        client = self.custodian_1_client
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        ds.record_queryset.delete()
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)

    def test_empty_not_allowed(self):
        ds = self.ds_1
        client = self.custodian_1_client
        payload = {
            "dataset": ds.pk,
            "data": {}
        }
        url = reverse('api:record-list')
        count = ds.record_queryset.count()
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), count)

    def test_create_column_not_in_schema(self):
        """
        Test that if we introduce a column not in the schema it will not validate in strict mode
        """
        ds = self.ds_1
        client = self.custodian_1_client
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31',
            'Extra Column': 'Extra Value'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = helpers.set_strict_mode(reverse('api:record-list'))
        ds.record_queryset.delete()
        self.assertEqual(
            client.post(url, data=payload, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), 0)

    def test_update_column_not_in_schema(self):
        """
        Test that updating a record with column not in the schema it will not validate in strict mode
        :return:
        """
        ds = self.ds_1
        client = self.custodian_1_client
        record = self._create_default_record()
        incorrect_data = clone(record.data)
        incorrect_data['Extra Column'] = "Extra Value"
        data = {
            "dataset": record.dataset.pk,
            "data": incorrect_data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        # set strict mode
        url = helpers.set_strict_mode(url)
        count = ds.record_queryset.count()
        self.assertEqual(
            client.put(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), count)
        self.assertEqual(
            client.patch(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), count)

    def test_date_error(self):
        """
        An observation must have a date
        :return:
        """
        ds = self.ds_1
        record = self._create_default_record()
        date_column = ds.schema.observation_date_field.name
        new_data = clone(record.data)
        url_post = reverse('api:record-list')
        url_update = reverse('api:record-detail', kwargs={'pk': record.pk})
        valid_values = ['15/08/2008']
        for value in valid_values:
            new_data[date_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }
            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_201_CREATED
            )
            self.assertEqual(ds.record_queryset.count(), count + 1)

        invalid_values = [None, '', 'not a date']
        for value in invalid_values:
            new_data[date_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }

            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.put(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.patch(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(ds.record_queryset.count(), count)

    def test_geometry_error(self):
        """
        An observation must have a valid geometry
        :return:
        """
        ds = self.ds_1
        record = self._create_default_record()
        lat_column = ds.schema.latitude_field.name
        new_data = clone(record.data)
        url_post = reverse('api:record-list')
        url_update = reverse('api:record-detail', kwargs={'pk': record.pk})
        valid_values = [-34.125]
        for value in valid_values:
            new_data[lat_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }
            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_201_CREATED
            )
            self.assertEqual(ds.record_queryset.count(), count + 1)

        invalid_values = [None, '', 'not a valid latitude']
        for value in invalid_values:
            new_data[lat_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }

            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.put(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.patch(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(ds.record_queryset.count(), count)

    def test_species_name(self):
        ds = self.ds_1
        record = self._create_default_record()
        column = ds.schema.species_name_parser.species_name_field.name
        new_data = clone(record.data)
        url_post = reverse('api:record-list')
        url_update = reverse('api:record-detail', kwargs={'pk': record.pk})
        valid_values = ['Canis Lupus', 'chubby bat', 'anything']
        for value in valid_values:
            new_data[column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }
            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_201_CREATED
            )
            self.assertEqual(ds.record_queryset.count(), count + 1)

        invalid_values = [None, '', 125]
        for value in invalid_values:
            new_data[column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }

            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.put(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.patch(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(ds.record_queryset.count(), count)


class TestDateTimeAndGeometryExtraction(helpers.BaseUserTestCase):
    species_facade_class = NoSpeciesFacade

    @staticmethod
    def schema_with_species_name():
        schema_fields = [
            {
                "name": "Species Name",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class

    def test_create(self):
        """
        Test that the date and geometry are extracted from the data
        and saved in DB
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }

        # clear all records
        ds.record_queryset.delete()
        self.assertEqual(ds.record_queryset.count(), 0)
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, data=payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        record = ds.record_queryset.first()
        expected_date = datetime.date(2018, 1, 31)
        self.assertEqual(timezone.localtime(record.datetime).date(), expected_date)
        geometry = record.geometry
        self.assertIsInstance(geometry, Point)
        self.assertEqual((115.75, -32.0), (geometry.x, geometry.y))

    def test_update(self):
        """
        Test that the date and geometry are extracted from the data
        and saved in DB after a PATCH of the record data
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }

        # clear all records
        ds.record_queryset.delete()
        self.assertEqual(ds.record_queryset.count(), 0)
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, data=payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        record = ds.record_queryset.first()
        # date and lat/lon
        # change lat/lon
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': 22.222,
            'Longitude': 111.111,
            'When': '2017-12-24'
        }
        payload = {
            "data": data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        self.assertEqual(
            client.patch(url, data=payload, format='json').status_code,
            status.HTTP_200_OK
        )
        record.refresh_from_db()
        expected_date = datetime.date(2017, 12, 24)
        self.assertEqual(timezone.localtime(record.datetime).date(), expected_date)
        geometry = record.geometry
        self.assertIsInstance(geometry, Point)
        self.assertEqual((111.111, 22.222), (geometry.x, geometry.y))


class TestSpeciesNameExtraction(helpers.BaseUserTestCase):
    species_facade_class = NoSpeciesFacade

    @staticmethod
    def schema_with_species_name():
        schema_fields = [
            {
                "name": "Species Name",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class

    def test_create(self):
        """
        Test that the species name is extracted from the data and saved in DB even if the species is not valid
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }

        # clear all records
        ds.record_queryset.delete()
        self.assertEqual(ds.record_queryset.count(), 0)
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, data=payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        self.assertEqual(ds.record_queryset.first().species_name, 'Chubby Bat')

    def test_update(self):
        """
        Test that name extraction after a PUT method
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        record = ds.record_queryset.first()
        self.assertEqual(record.species_name, 'Chubby Bat')

        # update the species_name
        data = {
            'Species Name': ' Canis lupus ',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        self.assertEqual(
            client.put(url, payload, format='json').status_code,
            status.HTTP_200_OK
        )
        record.refresh_from_db()
        self.assertEqual(record.species_name, 'Canis lupus')

    def test_patch(self):
        """
        Test that name extraction after a PATCH method
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        record = ds.record_queryset.first()
        self.assertEqual(record.species_name, 'Chubby Bat')

        # update the species_name
        data = {
            'Species Name': 'Canis lupus ',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "data": data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        self.assertEqual(
            client.patch(url, payload, format='json').status_code,
            status.HTTP_200_OK
        )
        record.refresh_from_db()
        self.assertEqual(record.species_name, 'Canis lupus')


class TestNameIDFromSpeciesName(helpers.BaseUserTestCase):
    """
    Test that we retrieve the name id from the species facade
    """

    species_facade_class = helpers.LightSpeciesFacade

    @staticmethod
    def schema_with_species_name():
        schema_fields = [
            {
                "name": "Species Name",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class

    def test_create(self):
        """
        Test that the name_id is retrieved from the species facade from the species_name
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        data = {
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        for species_name, name_id in list(helpers.LightSpeciesFacade().name_id_by_species_name().items())[:2]:
            ds.record_queryset.delete()
            self.assertEqual(ds.record_queryset.count(), 0)
            data['Species Name'] = species_name
            payload = {
                "dataset": ds.pk,
                "data": data
            }
            url = reverse('api:record-list')
            self.assertEqual(
                client.post(url, payload, format='json').status_code,
                status.HTTP_201_CREATED
            )
            self.assertEqual(ds.record_queryset.count(), 1)
            self.assertEqual(ds.record_queryset.first().name_id, name_id)

    def test_update(self):
        """
        Test that the name_id is retrieved from the species facade from the species_name
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        # create a record with a wrong species name. Should have name_id = -1
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        record = ds.record_queryset.first()
        self.assertEqual(record.name_id, -1)

        # update the species_name
        data = {
            'Species Name': 'Canis lupus',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        expected_name_id = 25454
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        self.assertEqual(
            client.put(url, payload, format='json').status_code,
            status.HTTP_200_OK
        )
        record.refresh_from_db()
        self.assertEqual(record.name_id, expected_name_id)

    def test_patch(self):
        """
        Same as above but wit a patch method instead of put
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_species_name()
        ds = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        # create a record with a wrong species name. Should have name_id = -1
        data = {
            'Species Name': 'Chubby Bat',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "dataset": ds.pk,
            "data": data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        record = ds.record_queryset.first()
        self.assertEqual(record.name_id, -1)

        # update the species_name
        data = {
            'Species Name': 'Canis lupus',
            'Latitude': -32.0,
            'Longitude': 115.75,
            'When': '2018-01-31'
        }
        payload = {
            "data": data
        }
        expected_name_id = 25454
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        self.assertEqual(
            client.patch(url, payload, format='json').status_code,
            status.HTTP_200_OK
        )
        record.refresh_from_db()
        self.assertEqual(record.name_id, expected_name_id)


class TestExport(helpers.BaseUserTestCase):

    def setUp(self):
        super(TestExport, self).setUp()
        rows = [
            ['When', 'Species Name', 'How Many', 'Latitude', 'Longitude', 'Comments'],
            ['2018-02-07', 'Canis lupus', 1, -32.0, 115.75, ''],
            ['2018-01-12', 'Chubby bat', 10, -32.0, 115.75, 'Awesome'],
            ['2018-02-02', 'Canis dingo', 2, -32.0, 115.75, 'Watch out kids'],
            ['2018-02-10', 'Unknown', 3, -32.0, 115.75, 'Canis?'],
        ]
        self.ds_1 = self._create_dataset_and_records_from_rows(rows)
        self.assertEqual(self.ds_1.type, Dataset.TYPE_SPECIES_OBSERVATION)

    def test_happy_path_no_filter(self):
        client = self.custodian_1_client
        dataset = self.ds_1
        all_records = Record.objects.filter(dataset=dataset)
        self.assertTrue(all_records.count() > 0)
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        # check headers
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=DatasetName_YYYY_MM_DD-HHMMSS.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEqual(ext, '.xlsx')
        filename.startswith(dataset.name)
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named from dataset
        sheet_names = wb.sheetnames
        self.assertEqual(1, len(sheet_names))
        self.assertEqual(dataset.name, sheet_names[0])
        ws = wb[dataset.name]
        rows = list(ws.rows)
        expected_records = Record.objects.filter(dataset=dataset)
        self.assertEqual(len(rows), expected_records.count() + 1)
        headers = [c.value for c in rows[0]]
        schema = dataset.schema
        # all the columns of the schema should be in the excel
        self.assertEqual(schema.headers, headers)

    def test_permission_ok_for_not_custodian(self):
        """Export is a read action. Should be authorised for every logged-in user."""
        client = self.custodian_2_client
        dataset = self.ds_1
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_permission_denied_if_not_logged_in(self):
        """Must be logged-in."""
        client = self.anonymous_client
        dataset = self.ds_1
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)


class TestSpeciesNameFromNameID(helpers.BaseUserTestCase):
    """
    Use case:
    The schema doesn't include a Species Name but just a Name Id column.
    Test that using the upload (excel) or API the species name is collected from herbie and populated.
    The test suite uses a mock herbie facade with a static species_name -> nameId dict
    @see helpers.SOME_SPECIES_NAME_NAME_ID_MAP
    """

    species_facade_class = helpers.LightSpeciesFacade

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class

    @staticmethod
    def schema_with_name_id():
        schema_fields = [
            {
                "name": "Name Id",
                "type": "integer",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def test_species_name_collected_upload(self):
        """
        Happy path: upload excel with a valid Name Id.
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        # data
        csv_data = [
            ['Name Id', 'When', 'Latitude', 'Longitude'],
            [25454, '01/01/2017', -32.0, 115.75],  # "Canis lupus"
            ['24204', '02/02/2017', -33.0, 116.0]  # "Vespadelus douglasorum"
        ]
        file_ = helpers.rows_to_xlsx_file(csv_data)
        self.assertEqual(0, Record.objects.filter(dataset=dataset).count())
        url = reverse('api:dataset-upload', kwargs={'pk': dataset.pk})
        with open(file_, 'rb') as fp:
            payload = {
                'file': fp
            }
            resp = client.post(url, payload, format='multipart')
            self.assertEqual(status.HTTP_200_OK, resp.status_code)
            records = Record.objects.filter(dataset=dataset)
            self.assertEqual(records.count(), len(csv_data) - 1)
            for r in records:
                self.assertTrue(r.name_id > 0)
                self.assertIsNotNone(r.species_name)
            canis_lupus = records.filter(name_id=25454).first()
            self.assertIsNotNone(canis_lupus)
            self.assertEqual(canis_lupus.species_name, "Canis lupus")
            vespadelus = records.filter(name_id=24204).first()
            self.assertIsNotNone(vespadelus)
            self.assertEqual(vespadelus.species_name, "Vespadelus douglasorum")

    def test_species_name_collected_api_create(self):
        """
        Same as above: testing that the species name is collected when using the API create
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        record_data = {
            'Name Id': 25454,  # "Canis lupus"
            'When': '12/12/2017',
            'Latitude': -32.0,
            'Longitude': 115.756
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.name_id, 25454)
        self.assertEqual(record.species_name, "Canis lupus")

    def test_species_name_collected_api_update(self):
        """
        Updating the Name Id should update the species name
        :return:
        """
        # create record
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        record_data = {
            'Name Id': 25454,  # "Canis lupus"
            'When': '12/12/2017',
            'Latitude': -32.0,
            'Longitude': 115.756
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.name_id, 25454)
        self.assertEqual(record.species_name, "Canis lupus")

        # patch Name Id
        new_name_id = 24204
        record_data['Name Id'] = new_name_id
        expected_species_name = 'Vespadelus douglasorum'
        url = reverse('api:record-detail', kwargs={'pk': record.pk})
        payload = {
            'data': record_data
        }
        resp = client.patch(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertEqual(record.name_id, new_name_id)
        self.assertEqual(record.species_name, expected_species_name)

    def test_wrong_id_rejected_upload(self):
        """
        If a wrong Name Id is provided the system assume its an error
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        # data
        csv_data = [
            ['Name Id', 'When', 'Latitude', 'Longitude'],
            [99934, '01/01/2017', -32.0, 115.75],  # wrong
            ['24204', '02/02/2017', -33.0, 116.0]  # "Vespadelus douglasorum"
        ]
        file_ = helpers.rows_to_xlsx_file(csv_data)
        self.assertEqual(0, Record.objects.filter(dataset=dataset).count())
        url = reverse('api:dataset-upload', kwargs={'pk': dataset.pk})
        with open(file_, 'rb') as fp:
            payload = {
                'file': fp
            }
            resp = client.post(url, payload, format='multipart')
            self.assertEqual(status.HTTP_400_BAD_REQUEST, resp.status_code)
            records = Record.objects.filter(dataset=dataset)
            # should be only one record (the good one)
            self.assertEqual(records.count(), 1)
            vespadelus = records.filter(name_id=24204).first()
            self.assertIsNotNone(vespadelus)
            self.assertEqual(vespadelus.species_name, "Vespadelus douglasorum")

    def test_wrong_id_rejected_api_create(self):
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        record_data = {
            'Name Id': 9999,  # wrong
            'When': '12/12/2017',
            'Latitude': -32.0,
            'Longitude': 115.756
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(Record.objects.filter(dataset=dataset).count(), 0)


class TestSpeciesNameAndNameID(helpers.BaseUserTestCase):
    """
    Use case:
    The schema includes a Species Name and a Name Id column.
    Test that the Name Id takes precedence
    The test suite uses a mock herbie facade with a static species_name -> Name Id dict
    @see helpers.SOME_SPECIES_NAME_NAME_ID_MAP
    """

    species_facade_class = helpers.LightSpeciesFacade

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class

    @staticmethod
    def schema_with_name_id_and_species_name():
        schema_fields = [
            {
                "name": "Name Id",
                "type": "integer",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "Species Name",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def test_species_name_collected_upload(self):
        """
        Happy path: upload excel with a valid Name Id.
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id_and_species_name()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        # data
        csv_data = [
            ['Name Id', 'Species Name', 'When', 'Latitude', 'Longitude'],
            [25454, 'Chubby Bat', '01/01/2017', -32.0, 115.75],  # "Canis lupus"
            ['24204', 'French Frog', '02/02/2017', -33.0, 116.0]  # "Vespadelus douglasorum"
        ]
        file_ = helpers.rows_to_xlsx_file(csv_data)
        self.assertEqual(0, Record.objects.filter(dataset=dataset).count())
        url = reverse('api:dataset-upload', kwargs={'pk': dataset.pk})
        with open(file_, 'rb') as fp:
            payload = {
                'file': fp
            }
            resp = client.post(url, payload, format='multipart')
            self.assertEqual(status.HTTP_200_OK, resp.status_code)
            records = Record.objects.filter(dataset=dataset)
            self.assertEqual(records.count(), len(csv_data) - 1)
            for r in records:
                self.assertTrue(r.name_id > 0)
                self.assertIsNotNone(r.species_name)
            canis_lupus = records.filter(name_id=25454).first()
            self.assertIsNotNone(canis_lupus)
            self.assertEqual(canis_lupus.species_name, "Canis lupus")
            vespadelus = records.filter(name_id=24204).first()
            self.assertIsNotNone(vespadelus)
            self.assertEqual(vespadelus.species_name, "Vespadelus douglasorum")

    def test_nameId_collected_upload(self):
        """
        Test that if Name Id is not provided it is collected from the species list
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id_and_species_name()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        # data
        csv_data = [
            ['Name Id', 'Species Name', 'When', 'Latitude', 'Longitude'],
            ['', 'Canis lupus', '01/01/2017', -32.0, 115.75],  # "Canis lupus"
            ['', 'Vespadelus douglasorum', '02/02/2017', -33.0, 116.0]  # "Vespadelus douglasorum"
        ]
        file_ = helpers.rows_to_xlsx_file(csv_data)
        self.assertEqual(0, Record.objects.filter(dataset=dataset).count())
        url = reverse('api:dataset-upload', kwargs={'pk': dataset.pk})
        with open(file_, 'rb') as fp:
            payload = {
                'file': fp
            }
            resp = client.post(url, payload, format='multipart')
            self.assertEqual(status.HTTP_200_OK, resp.status_code)
            records = Record.objects.filter(dataset=dataset)
            self.assertEqual(records.count(), len(csv_data) - 1)
            for r in records:
                self.assertTrue(r.name_id > 0)
                self.assertIsNotNone(r.species_name)
            canis_lupus = records.filter(name_id=25454).first()
            self.assertIsNotNone(canis_lupus)
            self.assertEqual(canis_lupus.species_name, "Canis lupus")
            vespadelus = records.filter(name_id=24204).first()
            self.assertIsNotNone(vespadelus)
            self.assertEqual(vespadelus.species_name, "Vespadelus douglasorum")

    def test_species_name_collected_api_create(self):
        """
        Same as above: testing that the species name is collected when using the API create
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id_and_species_name()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        record_data = {
            'Name Id': 25454,  # "Canis lupus"
            'Species Name': 'Chubby Bat',
            'When': '12/12/2017',
            'Latitude': -32.0,
            'Longitude': 115.756
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.name_id, 25454)
        self.assertEqual(record.species_name, "Canis lupus")

    def test_species_name_collected_api_update(self):
        """
        Updating the Name Id should update the species name
        :return:
        """
        # create record
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_name_id_and_species_name()
        dataset = self._create_dataset_with_schema(
            project, self.data_engineer_1_client, schema, dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
        )
        record_data = {
            'Name Id': 25454,  # "Canis lupus"
            'Species Name': 'Chubby Bat',
            'When': '12/12/2017',
            'Latitude': -32.0,
            'Longitude': 115.756
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.name_id, 25454)
        self.assertEqual(record.species_name, "Canis lupus")
        # TODO: the species name in the data is not updated. Should we?
        self.assertEqual(record.data.get('Species Name'), 'Chubby Bat')

        # patch Name Id
        new_name_id = 24204
        record_data['Name Id'] = new_name_id
        expected_species_name = 'Vespadelus douglasorum'
        url = reverse('api:record-detail', kwargs={'pk': record.pk})
        payload = {
            'data': record_data
        }
        resp = client.patch(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertEqual(record.name_id, new_name_id)
        self.assertEqual(record.species_name, expected_species_name)


class TestCompositeSpeciesName(helpers.BaseUserTestCase):
    """
    Test for species name composed from Genus, Species, infra_rank, infra_name columns
    """

    species_facade_class = helpers.LightSpeciesFacade

    @staticmethod
    def schema_with_4_columns_genus():
        schema_fields = [
            {
                "name": "Genus",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Species",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "InfraSpecific Rank",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "InfraSpecific Name",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    @staticmethod
    def schema_with_2_columns_genus():
        schema_fields = [
            {
                "name": "Genus",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Species",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    @staticmethod
    def schema_with_genus_and_species_name_no_required():
        schema_fields = [
            {
                "name": "SpeciesName",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": "speciesName"
                }
            },
            {
                "name": "Genus",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": "genus"
                }
            },
            {
                "name": "Species",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": "species"
                }
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def _more_setup(self):
        # set the HerbieFacade class
        from main.api.views import SpeciesMixin
        SpeciesMixin.species_facade_class = self.species_facade_class
        self.client = self.custodian_1_client

    def assert_create_dataset(self, schema):
        try:
            return self._create_dataset_with_schema(
                self.project_1,
                self.data_engineer_1_client,
                schema,
                dataset_type=Dataset.TYPE_SPECIES_OBSERVATION
            )
        except Exception as e:
            self.fail('Species Observation dataset creation failed for schema {schema}'.format(
                schema=schema
            ))

    def test_genus_species_only_happy_path(self):
        schema = self.schema_with_2_columns_genus()
        dataset = self.assert_create_dataset(schema)
        records = [
            ['Genus', 'Species', 'When', 'Latitude', 'Longitude'],
            ['Canis', 'lupus', '2018-01-25', -32.0, 115.75],
        ]
        resp = self._upload_records_from_rows(records, dataset_pk=dataset.pk)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        received = resp.json()
        rec_id = received[0]['recordId']
        record = Record.objects.filter(pk=rec_id).first()
        self.assertEqual(record.species_name, 'Canis lupus')
        self.assertEqual(record.name_id, 25454)

    def test_genus_species_and_infra_specifics_happy_path(self):
        schema = self.schema_with_4_columns_genus()
        dataset = self.assert_create_dataset(schema)
        records = [
            ['Genus', 'Species', 'InfraSpecific Rank', 'InfraSpecific Name', 'When', 'Latitude', 'Longitude'],
            ['Canis', 'lupus', 'subsp. familiaris ', ' rank naughty dog ', '2018-01-25', -32.0, 115.75],
        ]
        resp = self._upload_records_from_rows(records, dataset_pk=dataset.pk)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        received = resp.json()
        rec_id = received[0]['recordId']
        record = Record.objects.filter(pk=rec_id).first()
        expected_species_name = 'Canis lupus subsp. familiaris rank naughty dog'
        self.assertEqual(record.species_name, expected_species_name)
        self.assertEqual(record.name_id, -1)

    def test_validation_missing_species(self):
        schema = self.schema_with_2_columns_genus()
        dataset = self.assert_create_dataset(schema)
        data = {
            'Genus': "Canis",
            'When': '2018-01-25',
            'Latitude': -32.0,
            'Longitude': 115.75
        }
        url = helpers.url_post_record_strict()
        payload = {
            'dataset': dataset.pk,
            'data': data
        }
        resp = self.client.post(url, payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        received_json = resp.json()
        # should contain one error on the 'data' for the species field
        self.assertIn('data', received_json)
        errors = received_json.get('data')
        self.assertIsInstance(errors, list)
        self.assertEqual(len(errors), 1)
        error = errors[0]
        # should be "Species::msg"
        pattern = re.compile(r"^Species::(.+)$")
        self.assertTrue(pattern.match(error))

    def test_genus_required_error(self):
        """
        If genus is set to be required and not provided it should not throw an exception
        but return a 400 with a field error message
        see https://decbugs.com/view.php?id=6907 for details
        """
        schema = self.schema_with_2_columns_genus()
        dataset = self.assert_create_dataset(schema)
        # Genus is required
        self.assertTrue(dataset.schema.get_field_by_name('Genus').required)

        # provides 3 records with no Genus (row=2,3,4)
        records = [
            ['Genus', 'Species', 'When', 'Latitude', 'Longitude'],
            [None, 'lupus', '2018-01-25', -32.0, 115.75],
            ['', 'lupus', '2018-01-25', -32.0, 115.75],
            ['  ', 'lupus', '2018-01-25', -32.0, 115.75]
        ]
        resp = self._upload_records_from_rows(records, dataset_pk=dataset.pk, strict=False)
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        received = resp.json()
        # expected: array of report by row
        self.assertIsInstance(received, list)
        self.assertEqual(len(received), 3)
        # this what an report should look like
        expected_row_report = {
            'row': 3,
            'errors': {'Genus': 'Field "Genus" has constraint "required" which is not satisfied for value "None"'},
            'warnings': {}}
        for row_report in received:
            self.assertIn('errors', row_report)
            errors = row_report.get('errors')
            self.assertIn('Genus', errors)
            msg = errors.get('Genus')
            self.assertEqual(msg, expected_row_report['errors']['Genus'])

    def test_species_required_error(self):
        """
        If species (with genus) is set to be required and not provided it should not throw an exception
        but return a 400 with a field error message
        see https://decbugs.com/view.php?id=6907 for details
        """
        schema = self.schema_with_2_columns_genus()
        dataset = self.assert_create_dataset(schema)
        # Genus is required
        self.assertTrue(dataset.schema.get_field_by_name('Genus').required)

        # provides 3 records with no Species (row=2,3,4)
        records = [
            ['Genus', 'Species', 'When', 'Latitude', 'Longitude'],
            ['Canis', '', '2018-01-25', -32.0, 115.75],
            ['Canis', None, '2018-01-25', -32.0, 115.75],
            ['Canis', '   ', '2018-01-25', -32.0, 115.75]
        ]
        resp = self._upload_records_from_rows(records, dataset_pk=dataset.pk, strict=False)
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        received = resp.json()
        # expected: array of report by row
        self.assertIsInstance(received, list)
        self.assertEqual(len(received), 3)
        # this what an report should look like
        expected_row_report = {
            'row': 3,
            'errors': {'Species': 'Field "Species" has constraint "required" which is not satisfied for value "None"'},
            'warnings': {}}
        for row_report in received:
            self.assertIn('errors', row_report)
            errors = row_report.get('errors')
            self.assertIn('Species', errors)
            msg = errors.get('Species')
            self.assertEqual(msg, expected_row_report['errors']['Species'])

    def test_species_name_and_genus_requirement(self):
        """
        If the schema has speciesName and genus/species we should not impose any requirement
        User should be able to choose one or the other way to enter a species.
        """
        schema = self.schema_with_genus_and_species_name_no_required()
        self.assert_create_dataset(schema)

    def test_species_name_tag_precedence(self):
        """
        if the schema has Species Name and genus/species and the the Species Name column is biosys tagged as type
        speciesName it then has precedence over genus/species.
        @see https://youtrack.gaiaresources.com.au/youtrack/issue/BIOSYS-305
        Given I have a species observation dataset with fields |Genus|Species|Species Name|
        And the Species Name field is tagged with the Biosys type 'SpeciesName'
        And Genus and Species fields have no Biosys type
        When I enter |Pteropyus|vampyrus|Canis lupus|
        Then the species extracted should be Canis lupus and not Pteropyus vampyrus
        """
        schema = self.schema_with_genus_and_species_name_no_required()
        # remove biosys tag for Genus and Species
        for field in schema['fields']:
            if field['name'] in ['Genus', 'Species']:
                del field['biosys']
        dataset = self.assert_create_dataset(schema)
        records = [
            ['Genus', 'Species', 'SpeciesName', 'When', 'Latitude', 'Longitude'],
            ['Pteropyus', 'vampyrus', 'Canis lupus', '2018-01-25', -32.0, 115.75],
        ]
        expected_species_name = 'Canis lupus'

        resp = self._upload_records_from_rows(records, dataset_pk=dataset.pk, strict=False)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        received = resp.json()
        rec_id = received[0]['recordId']
        record = Record.objects.filter(pk=rec_id).first()
        self.assertEqual(record.species_name, expected_species_name)


class TestPatch(helpers.BaseUserTestCase):

    def test_patch_validated(self):
        """
        Test that we can patch just the 'validated' flag
        :return:
        """
        rows = [
            ['Species Name', 'When', 'Latitude', 'Longitude', 'Comments'],
            ['Chubby bat', '2018-06-01', -32, 115.75, 'It is huge!']
        ]
        dataset = self._create_dataset_and_records_from_rows(rows)
        self.assertEqual(dataset.type, Dataset.TYPE_SPECIES_OBSERVATION)
        records = dataset.record_set.all()
        record = records.last()
        self.assertIsNotNone(record)
        self.assertFalse(record.validated)
        previous_data = json.dumps(record.data)
        # patch
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        payload = {
            'validated': True
        }
        resp = client.patch(url, payload)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertTrue(record.validated)
        self.assertTrue(json.dumps(record.data), previous_data)

    def test_patch_locked(self):
        """
        Test that we can patch just the 'locked' flag
        :return:
        """
        rows = [
            ['Species Name', 'When', 'Latitude', 'Longitude', 'Comments'],
            ['Chubby bat', '2018-06-01', -32, 115.75, 'It is huge!']
        ]
        dataset = self._create_dataset_and_records_from_rows(rows)
        self.assertEqual(dataset.type, Dataset.TYPE_SPECIES_OBSERVATION)
        records = dataset.record_set.all()
        record = records.last()
        self.assertIsNotNone(record)
        self.assertFalse(record.locked)
        previous_data = json.dumps(record.data)
        # patch
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        payload = {
            'locked': True
        }
        resp = client.patch(url, payload)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertTrue(record.locked)
        self.assertTrue(json.dumps(record.data), previous_data)
