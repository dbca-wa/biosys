import datetime
import re
from os import path
import json

from django.contrib.auth.models import User
from django.contrib.gis.geos import Point
from django.core.urlresolvers import reverse
from django.test import TestCase, override_settings
from django.utils import timezone, six
# TODO: replace all the use of G by a factory from factory-boy
from django_dynamic_fixture import G
from openpyxl import load_workbook
from rest_framework import status
from rest_framework.test import APIClient

from main import constants
from main.models import Project, Site, Dataset, Record
from main.tests.api import helpers
from main.tests.test_data_package import clone
from main.utils_auth import is_admin
from main.tests import factories


class TestPermissions(TestCase):
    """
    Test Permissions
    Get: authenticated
    Update: admin, custodians
    Create: admin, custodians
    Delete: admin, custodians
    """
    fixtures = [
        'test-users',
        'test-projects',
        'test-sites',
        'test-datasets',
        'test-observations'
    ]

    @override_settings(PASSWORD_HASHERS=('django.contrib.auth.hashers.MD5PasswordHasher',),
                       REST_FRAMEWORK_TEST_SETTINGS=helpers.REST_FRAMEWORK_TEST_SETTINGS)
    def setUp(self):
        password = 'password'
        self.admin_user = User.objects.filter(username="admin").first()
        self.assertIsNotNone(self.admin_user)
        self.assertTrue(is_admin(self.admin_user))
        self.admin_user.set_password(password)
        self.admin_user.save()
        self.admin_client = APIClient()
        self.assertTrue(self.admin_client.login(username=self.admin_user.username, password=password))

        self.custodian_1_user = User.objects.filter(username="custodian1").first()
        self.assertIsNotNone(self.custodian_1_user)
        self.custodian_1_user.set_password(password)
        self.custodian_1_user.save()
        self.custodian_1_client = APIClient()
        self.assertTrue(self.custodian_1_client.login(username=self.custodian_1_user.username, password=password))
        self.project_1 = Project.objects.filter(name="Project1").first()
        self.site_1 = Site.objects.filter(code="Site1").first()
        self.ds_1 = Dataset.objects.filter(name="Observation1", project=self.project_1).first()
        self.assertIsNotNone(self.ds_1)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = self.ds_1.record_model.objects.filter(dataset=self.ds_1).first()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))

        self.custodian_2_user = User.objects.filter(username="custodian2").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.custodian_2_user.set_password(password)
        self.custodian_2_user.save()
        self.custodian_2_client = APIClient()
        self.assertTrue(self.custodian_2_client.login(username=self.custodian_2_user.username, password=password))
        self.project_2 = Project.objects.filter(name="Project2").first()
        self.site_2 = Site.objects.filter(code="Site2").first()
        self.ds_2 = Dataset.objects.filter(name="Observation2", project=self.project_2).first()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

        self.readonly_user = User.objects.filter(username="readonly").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.assertFalse(self.site_2.is_custodian(self.readonly_user))
        self.assertFalse(self.site_1.is_custodian(self.readonly_user))
        self.readonly_user.set_password(password)
        self.readonly_user.save()
        self.readonly_client = APIClient()
        self.assertTrue(self.readonly_client.login(username=self.readonly_user.username, password=password))

        self.anonymous_client = APIClient()

    def test_get(self):
        urls = [
            reverse('api:record-list'),
            reverse('api:record-detail', kwargs={'pk': 1})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.get(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.get(url).status_code,
                    status.HTTP_200_OK
                )

    def test_create(self):
        """
        Admin and custodians
        :return:
        """
        urls = [reverse('api:record-list')]
        ds = self.ds_1
        rec = self.record_1
        data = {
            "dataset": rec.dataset.pk,
            "site": rec.site.pk,
            "data": rec.data,
            "datetime": rec.datetime,
            "geometry": rec.geometry.geojson
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                count = ds.record_queryset.count()
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )
                self.assertEqual(ds.record_queryset.count(), count + 1)

    def test_bulk_create(self):
        """
        Cannot create bulk with this end point
        :return:
        """
        urls = [reverse('api:record-list')]
        rec = self.record_1
        ds = self.ds_1
        data = [
            {
                "dataset": rec.dataset.pk,
                "site": rec.site.pk,
                "data": rec.data
            },
            {
                "dataset": rec.dataset.pk,
                "site": rec.site.pk,
                "data": rec.data
            }
        ]
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client,
                          self.admin_client, self.custodian_1_client],
            "allowed": []
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_400_BAD_REQUEST, status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                count = ds.record_queryset.count()
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )
                self.assertEqual(ds.record_queryset.count(), count + len(data))

    def test_update(self):
        """
        admin + custodian of project for site 1
        :return:
        """
        rec = self.record_1
        previous_data = clone(rec.data)
        updated_data = clone(previous_data)
        updated_data['Longitude'] = '118.78'
        urls = [reverse('api:record-detail', kwargs={'pk': rec.pk})]
        data = {
            "data": updated_data,
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.patch(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                rec.data = previous_data
                rec.save()
                self.assertEqual(
                    client.patch(url, data, format='json').status_code,
                    status.HTTP_200_OK
                )
                rec.refresh_from_db()
                self.assertEqual(rec.data, updated_data)

    def test_delete(self):
        """
        Currently admin + custodian
        :return:
        """
        rec = self.record_1
        urls = [reverse('api:record-detail', kwargs={'pk': rec.pk})]
        data = None
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.delete(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                rec.save()
                count = Dataset.objects.count()
                self.assertEqual(
                    client.delete(url, data, format='json').status_code,
                    status.HTTP_204_NO_CONTENT
                )
                self.assertTrue(Dataset.objects.count(), count - 1)

    def test_options(self):
        urls = [
            reverse('api:record-list'),
            reverse('api:record-detail', kwargs={'pk': 1})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.options(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        # authenticated
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.options(url).status_code,
                    status.HTTP_200_OK
                )


class TestDataValidation(TestCase):
    fixtures = [
        'test-users',
        'test-projects',
        'test-sites',
        'test-datasets',
        'test-observations'
    ]

    @override_settings(PASSWORD_HASHERS=('django.contrib.auth.hashers.MD5PasswordHasher',),
                       REST_FRAMEWORK_TEST_SETTINGS=helpers.REST_FRAMEWORK_TEST_SETTINGS)
    def setUp(self):
        password = 'password'
        self.admin_user = User.objects.filter(username="admin").first()
        self.assertIsNotNone(self.admin_user)
        self.assertTrue(is_admin(self.admin_user))
        self.admin_user.set_password(password)
        self.admin_user.save()
        self.admin_client = APIClient()
        self.assertTrue(self.admin_client.login(username=self.admin_user.username, password=password))

        self.custodian_1_user = User.objects.filter(username="custodian1").first()
        self.assertIsNotNone(self.custodian_1_user)
        self.custodian_1_user.set_password(password)
        self.custodian_1_user.save()
        self.custodian_1_client = APIClient()
        self.assertTrue(self.custodian_1_client.login(username=self.custodian_1_user.username, password=password))
        self.project_1 = Project.objects.filter(name="Project1").first()
        self.site_1 = Site.objects.filter(code="Site1").first()
        self.ds_1 = Dataset.objects.filter(name="Observation1", project=self.project_1).first()
        self.assertIsNotNone(self.ds_1)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = self.ds_1.record_model.objects.filter(dataset=self.ds_1).first()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))

        self.custodian_2_user = User.objects.filter(username="custodian2").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.custodian_2_user.set_password(password)
        self.custodian_2_user.save()
        self.custodian_2_client = APIClient()
        self.assertTrue(self.custodian_2_client.login(username=self.custodian_2_user.username, password=password))
        self.project_2 = Project.objects.filter(name="Project2").first()
        self.site_2 = Site.objects.filter(code="Site2").first()
        self.ds_2 = Dataset.objects.filter(name="Observation2", project=self.project_2).first()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

        self.readonly_user = User.objects.filter(username="readonly").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.assertFalse(self.site_2.is_custodian(self.readonly_user))
        self.assertFalse(self.site_1.is_custodian(self.readonly_user))
        self.readonly_user.set_password(password)
        self.readonly_user.save()
        self.readonly_client = APIClient()
        self.assertTrue(self.readonly_client.login(username=self.readonly_user.username, password=password))

        self.anonymous_client = APIClient()

    def test_create_one_happy_path(self):
        """
        Test the create of one record
        :return:
        """
        # grab one existing an re-inject it
        record = self.record_1
        ds = self.ds_1
        data = {
            "dataset": record.dataset.pk,
            "data": record.data
        }
        url = reverse('api:record-list')
        client = self.custodian_1_client
        count = ds.record_queryset.count()
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), count + 1)

    def test_empty_not_allowed(self):
        ds = self.ds_1
        record = self.record_1
        data = {
            "dataset": record.dataset.pk,
            "data": {}
        }
        url = reverse('api:record-list')
        client = self.custodian_1_client
        count = ds.record_queryset.count()
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), count)

    def test_create_column_not_in_schema(self):
        """
        Test that if we introduce a column not in the the dataset it will not validate
        :return:
        """
        ds = self.ds_1
        record = self.record_1
        incorrect_data = clone(record.data)
        incorrect_data['Extra Column'] = "Extra Value"
        data = {
            "dataset": record.dataset.pk,
            "data": incorrect_data
        }
        url = reverse('api:record-list')
        # set strict mode
        url = helpers.set_strict_mode(url)
        client = self.custodian_1_client
        count = ds.record_queryset.count()
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), count)

    def test_update_column_not_in_schema(self):
        """
        Test that if we introduce a column not in the the dataset it will not validate
        :return:
        """
        ds = self.ds_1
        record = self.record_1
        incorrect_data = clone(record.data)
        incorrect_data['Extra Column'] = "Extra Value"
        data = {
            "dataset": record.dataset.pk,
            "data": incorrect_data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        count = ds.record_queryset.count()
        # set strict mode
        url = helpers.set_strict_mode(url)
        self.assertEqual(
            client.put(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), count)
        self.assertEqual(
            client.patch(url, data, format='json').status_code,
            status.HTTP_400_BAD_REQUEST
        )
        self.assertEqual(ds.record_queryset.count(), count)

    def test_date_error(self):
        """
        An observation must have a date
        :return:
        """
        ds = self.ds_1
        record = self.record_1
        date_column = ds.schema.observation_date_field.name
        new_data = clone(record.data)
        url_post = reverse('api:record-list')
        url_update = reverse('api:record-detail', kwargs={'pk': record.pk})
        valid_values = ['15/08/2008']
        for value in valid_values:
            new_data[date_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }
            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_201_CREATED
            )
            self.assertEqual(ds.record_queryset.count(), count + 1)

        invalid_values = [None, '', 'abcd']
        for value in invalid_values:
            new_data[date_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }

            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.put(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.patch(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(ds.record_queryset.count(), count)

    def test_geometry_error(self):
        """
        An observation must have a valid geometry
        :return:
        """
        ds = self.ds_1
        record = self.record_1
        lat_column = ds.schema.latitude_field.name
        new_data = clone(record.data)
        url_post = reverse('api:record-list')
        url_update = reverse('api:record-detail', kwargs={'pk': record.pk})
        valid_values = [-34.125]
        for value in valid_values:
            new_data[lat_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }
            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_201_CREATED
            )
            self.assertEqual(ds.record_queryset.count(), count + 1)

        invalid_values = [None, '', 'abcd']
        for value in invalid_values:
            new_data[lat_column] = value
            data = {
                "dataset": record.dataset.pk,
                "data": new_data
            }

            client = self.custodian_1_client
            count = ds.record_queryset.count()
            self.assertEqual(
                client.post(url_post, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.put(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(
                client.patch(url_update, data, format='json').status_code,
                status.HTTP_400_BAD_REQUEST
            )
            self.assertEqual(ds.record_queryset.count(), count)


class TestSiteExtraction(TestCase):
    fixtures = [
        'test-users',
        'test-projects',
        'test-sites',
        'test-datasets',
        'test-observations'
    ]

    @override_settings(PASSWORD_HASHERS=('django.contrib.auth.hashers.MD5PasswordHasher',),
                       REST_FRAMEWORK_TEST_SETTINGS=helpers.REST_FRAMEWORK_TEST_SETTINGS)
    def setUp(self):
        password = 'password'
        self.admin_user = User.objects.filter(username="admin").first()
        self.assertIsNotNone(self.admin_user)
        self.assertTrue(is_admin(self.admin_user))
        self.admin_user.set_password(password)
        self.admin_user.save()
        self.admin_client = APIClient()
        self.assertTrue(self.admin_client.login(username=self.admin_user.username, password=password))

        self.custodian_1_user = User.objects.filter(username="custodian1").first()
        self.assertIsNotNone(self.custodian_1_user)
        self.custodian_1_user.set_password(password)
        self.custodian_1_user.save()
        self.custodian_1_client = APIClient()
        self.assertTrue(self.custodian_1_client.login(username=self.custodian_1_user.username, password=password))
        self.project_1 = Project.objects.filter(name="Project1").first()
        self.site_1 = Site.objects.filter(code="Adolphus").first()
        self.ds_1 = Dataset.objects.filter(name="Observation1", project=self.project_1).first()
        self.assertIsNotNone(self.ds_1)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = self.ds_1.record_model.objects.filter(dataset=self.ds_1).first()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))
        self.assertTrue(self.record_1.site, self.site_1)

        self.custodian_2_user = User.objects.filter(username="custodian2").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.custodian_2_user.set_password(password)
        self.custodian_2_user.save()
        self.custodian_2_client = APIClient()
        self.assertTrue(self.custodian_2_client.login(username=self.custodian_2_user.username, password=password))
        self.project_2 = Project.objects.filter(name="Project2").first()
        self.site_2 = Site.objects.filter(code="Site2").first()
        self.ds_2 = Dataset.objects.filter(name="Observation2", project=self.project_2).first()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

        self.readonly_user = User.objects.filter(username="readonly").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.assertFalse(self.site_2.is_custodian(self.readonly_user))
        self.assertFalse(self.site_1.is_custodian(self.readonly_user))
        self.readonly_user.set_password(password)
        self.readonly_user.save()
        self.readonly_client = APIClient()
        self.assertTrue(self.readonly_client.login(username=self.readonly_user.username, password=password))

        self.anonymous_client = APIClient()

    def test_create_with_site(self):
        """
        The descriptor contains a foreign key to the site.
        Test that the site is extracted from the data
        :return:
        """
        # clear all records
        ds = self.ds_1
        ds.record_queryset.delete()
        self.assertEqual(ds.record_queryset.count(), 0)
        record = self.record_1
        data = {
            "dataset": record.dataset.pk,
            "data": record.data
        }
        schema = ds.schema
        self.assertTrue(schema.has_fk_for_model('Site'))
        expected_site = record.site
        url = reverse('api:record-list')
        client = self.custodian_1_client
        self.assertEqual(
            client.post(url, data, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(ds.record_queryset.count(), 1)
        self.assertEqual(ds.record_queryset.first().site, expected_site)

    def test_update_site(self):
        ds = self.ds_1
        record = ds.record_queryset.filter(site=self.site_1).first()
        self.assertIsNotNone(record)
        site = Site.objects.filter(name="Site1").first()
        # need to test if the site belongs to the dataset project or the update won't happen
        self.assertIsNotNone(site)
        self.assertEqual(site.project, record.dataset.project)
        self.assertNotEqual(record.site, site)
        # update site value
        schema = record.dataset.schema
        site_column = schema.get_fk_for_model('Site').data_field
        r_data = record.data
        r_data[site_column] = site.code
        data = {
            "data": r_data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        self.assertEqual(
            client.patch(url, data, format='json').status_code,
            status.HTTP_200_OK
        )
        record.refresh_from_db()
        self.assertEqual(record.site, site)


class TestDateTimeAndGeometryExtraction(helpers.BaseUserTestCase):

    @staticmethod
    def schema_with_lat_long_and_date():
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "biosys": {
                    "type": "latitude"
                },
                "constraints": {
                    "required": True,
                    "minimum": -90.0,
                    "maximum": 90.0,
                }
            },
            {
                "name": "Longitude",
                "type": "number",
                "biosys": {
                    "type": "longitude"
                },
                "constraints": {
                    "required": True,
                    "minimum": -180.0,
                    "maximum": 180.0,
                }
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    @staticmethod
    def schema_with_no_date():
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Latitude",
                "type": "number",
                "biosys": {
                    "type": "latitude"
                },
                "constraints": {
                    "required": True,
                    "minimum": -90.0,
                    "maximum": 90.0,
                }
            },
            {
                "name": "Longitude",
                "type": "number",
                "biosys": {
                    "type": "longitude"
                },
                "constraints": {
                    "required": True,
                    "minimum": -180.0,
                    "maximum": 180.0,
                }
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        return schema

    def test_create(self):
        """
        Test that the date and geometry are extracted from the data
        and saved in DB
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_lat_long_and_date()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertEqual(dataset.record_queryset.count(), 0)
        record_data = {
            'What': 'A test',
            'When': '01/06/2017',
            'Latitude': -32.0,
            'Longitude': 116.0
        }
        payload = {
            "dataset": dataset.pk,
            "data": record_data
        }
        expected_date = datetime.date(2017, 6, 1)
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(dataset.record_queryset.count(), 1)
        record = dataset.record_queryset.first()
        self.assertEqual(timezone.localtime(record.datetime).date(), expected_date)
        geometry = record.geometry
        self.assertIsInstance(geometry, Point)
        self.assertEqual(geometry.x, 116.0)
        self.assertEqual(geometry.y, -32.0)

    def test_update(self):
        """
        Test that the date and geometry are extracted from the data
        and saved in DB
        :return:
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_lat_long_and_date()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertEqual(dataset.record_queryset.count(), 0)
        record_data = {
            'What': 'A test',
            'When': '01/06/2017',
            'Latitude': -32.0,
            'Longitude': 116.0
        }
        payload = {
            "dataset": dataset.pk,
            "data": record_data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        record = dataset.record_queryset.first()

        # change date
        new_date = '20/4/2016'
        # change lat/lon
        new_long = 111.111
        new_lat = 22.222

        record_data = {
            'When': new_date,
            'Latitude': new_lat,
            'Longitude': new_long
        }
        payload = {
            "dataset": dataset.pk,
            "data": record_data
        }

        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        self.assertEqual(
            client.patch(url, data=payload, format='json').status_code,
            status.HTTP_200_OK
        )
        self.assertEqual(dataset.record_queryset.count(), 1)
        record.refresh_from_db()
        expected_date = datetime.date(2016, 4, 20)
        self.assertEqual(timezone.localtime(record.datetime).date(), expected_date)
        geometry = record.geometry
        self.assertIsInstance(geometry, Point)
        self.assertEqual(geometry.x, new_long)
        self.assertEqual(geometry.y, new_lat)

    def test_create_without_date(self):
        """
        As of 29/06/2017. Date are not mandatory to create a Observation record
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_no_date()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertEqual(dataset.record_queryset.count(), 0)
        record_data = {
            'What': 'A test',
            'Latitude': -32.0,
            'Longitude': 116.0
        }
        payload = {
            "dataset": dataset.pk,
            "data": record_data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        self.assertEqual(dataset.record_queryset.count(), 1)
        record = dataset.record_queryset.first()
        self.assertIsNone(record.datetime)
        geometry = record.geometry
        self.assertIsInstance(geometry, Point)
        self.assertEqual(geometry.x, 116.0)
        self.assertEqual(geometry.y, -32.0)

    def test_update_without_date(self):
        """
        As of 29/06/2017. Date are not mandatory to create a Observation record
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_no_date()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertEqual(dataset.record_queryset.count(), 0)
        record_data = {
            'What': 'A test',
            'Latitude': -32.0,
            'Longitude': 116.0
        }
        payload = {
            "dataset": dataset.pk,
            "data": record_data
        }
        url = reverse('api:record-list')
        self.assertEqual(
            client.post(url, payload, format='json').status_code,
            status.HTTP_201_CREATED
        )
        record = dataset.record_queryset.first()

        new_long = 111.111
        new_lat = 22.222
        record_data = {
            'Latitude': new_lat,
            'Longitude': new_long
        }
        payload = {
            "dataset": dataset.pk,
            "data": record_data
        }
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        self.assertEqual(
            client.patch(url, data=payload, format='json').status_code,
            status.HTTP_200_OK
        )
        self.assertEqual(dataset.record_queryset.count(), 1)
        record.refresh_from_db()
        self.assertIsNone(record.datetime)
        geometry = record.geometry
        self.assertIsInstance(geometry, Point)
        self.assertEqual(geometry.x, new_long)
        self.assertEqual(geometry.y, new_lat)


class TestEastingNorthing(helpers.BaseUserTestCase):
    """
    Use case: the schema contains a datum and a zone field and easting/northing.
    """

    @staticmethod
    def schema_with_easting_northing():
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Northing",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": "northing"
                }
            },
            {
                "name": "Easting",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": "easting"
                }
            },
            {
                "name": "Datum",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "Zone",
                "type": "integer",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            }
        ]
        return helpers.create_schema_from_fields(schema_fields)

    def test_create_happy_path(self):
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_easting_northing()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.datum_field)
        self.assertIsNotNone(dataset.schema.zone_field)

        easting = 405542.537
        northing = 6459127.469
        datum = 'GDA94'
        zone = 50
        record_data = {
            'What': 'Chubby Bat',
            'When': '12/12/2017',
            'Easting': easting,
            'Northing': northing,
            'Datum': datum,
            'Zone': zone
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        qs = dataset.record_queryset
        self.assertEqual(qs.count(), 1)
        record = qs.first()
        geom = record.geometry
        # should be in WGS84 -> srid = 4326
        self.assertEqual(geom.srid, 4326)
        # convert it back to GAD / zone 50 -> srid = 28350
        geom.transform(28350)
        # compare with 2 decimal place precision
        self.assertAlmostEqual(geom.x, easting, places=2)
        self.assertAlmostEqual(geom.y, northing, places=2)

    def test_update_happy_path(self):
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_easting_northing()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.datum_field)
        self.assertIsNotNone(dataset.schema.zone_field)

        # first create record with wrong zone
        easting = 405542.537
        northing = 6459127.469
        datum = 'GDA94'
        zone = 58
        record_data = {
            'What': 'Chubby Bat',
            'When': '12/12/2017',
            'Easting': easting,
            'Northing': northing,
            'Datum': datum,
            'Zone': zone
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        qs = dataset.record_queryset
        self.assertEqual(qs.count(), 1)
        record = qs.first()
        geom = record.geometry
        # should be in WGS84 -> srid = 4326
        self.assertEqual(geom.srid, 4326)
        # convert it back to GAD / zone 50 -> srid = 28350
        geom.transform(28350)
        # compare with 2 decimal place precision. Should be different that of expected
        self.assertNotAlmostEqual(geom.x, easting, places=2)
        self.assertNotAlmostEqual(geom.y, northing, places=2)

        # send path to update the zone
        record_data = {
            'What': 'Chubby Bat',
            'When': '12/12/2017',
            'Easting': easting,
            'Northing': northing,
            'Datum': datum,
            'Zone': 50
        }
        payload = {
            'data': record_data
        }
        url = reverse('api:record-detail', kwargs={'pk': record.pk})
        resp = client.patch(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        geom = record.geometry
        # should be in WGS84 -> srid = 4326
        self.assertEqual(geom.srid, 4326)
        # convert it back to GAD / zone 50 -> srid = 28350
        geom.transform(28350)
        self.assertAlmostEqual(geom.x, easting, places=2)
        self.assertAlmostEqual(geom.y, northing, places=2)

    def test_default_datum(self):
        """
        If only easting and northing are provided the project's datum/zone should be used
        """
        project = self.project_1
        srid = constants.get_datum_srid('GDA94 / MGA zone 50')
        self.assertEqual(srid, 28350)
        project.datum = srid
        project.save()
        client = self.custodian_1_client
        # schema with datum and zone not required
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Northing",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": "northing"
                }
            },
            {
                "name": "Easting",
                "type": "number",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": "easting"
                }
            },
            {
                "name": "Datum",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
            {
                "name": "Zone",
                "type": "integer",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            }
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.datum_field)
        self.assertIsNotNone(dataset.schema.zone_field)

        easting = 405542.537
        northing = 6459127.469
        record_data = {
            'What': 'Chubby Bat',
            'When': '12/12/2017',
            'Easting': easting,
            'Northing': northing,
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list') + '?strict=true'
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        qs = dataset.record_queryset
        self.assertEqual(qs.count(), 1)
        record = qs.first()
        geom = record.geometry
        # should be in WGS84 -> srid = 4326
        self.assertEqual(geom.srid, 4326)
        self.assertIsInstance(geom, Point)
        self.assertAlmostEqual(geom.x, 116, places=2)
        self.assertAlmostEqual(geom.y, -32, places=2)
        # convert it back to GAD / zone 50 -> srid = 28350
        geom.transform(srid)
        # compare with 2 decimal place precision
        self.assertAlmostEqual(geom.x, easting, places=2)
        self.assertAlmostEqual(geom.y, northing, places=2)


class TestGeometryFromSite(helpers.BaseUserTestCase):
    """
     Use case: the observation dataset doesn't contain any geometry columns/fields
     but a reference (foreign key) to the site code. In this case the when yhe user uploads observations with a site
     reference only the observation geometry should be copied (not referenced) from the site geometry.
    """

    @staticmethod
    def schema_with_site_code_fk():
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Site Code",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        schema = helpers.add_model_field_foreign_key_to_schema(schema, {
            'schema_field': 'Site Code',
            'model': 'Site',
            'model_field': 'code'
        })
        return schema

    @staticmethod
    def schema_with_latlong_and_site_code_fk():
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "format": "any",
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Latitude",
                "type": "number",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": 'latitude'
                }
            },
            {
                "name": "Longitude",
                "type": "number",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
                "biosys": {
                    "type": 'longitude'
                }
            },
            {
                "name": "Site Code",
                "type": "string",
                "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        schema = helpers.add_model_field_foreign_key_to_schema(schema, {
            'schema_field': 'Site Code',
            'model': 'Site',
            'model_field': 'code'
        })
        return schema

    def test_observation_schema_valid_with_site_foreign_key(self):
        """
        An observation schema should be valid without geometry fields as long it has a foreign key to site.
        """
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Site Code",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        schema = helpers.add_model_field_foreign_key_to_schema(schema, {
            'schema_field': 'Site Code',
            'model': 'Site',
            'model_field': 'code'
        })
        data_package = helpers.create_data_package_from_schema(schema)
        # create data set
        url = reverse('api:dataset-list')
        project = self.project_1
        client = self.custodian_1_client
        dataset_name = "Observation with site foreign key and no geometry"
        payload = {
            "name": dataset_name,
            "type": Dataset.TYPE_OBSERVATION,
            "project": project.pk,
            'data_package': data_package
        }

        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        # double check
        self.assertIsNotNone(Dataset.objects.filter(project=project, name=dataset_name).first())

    def test_observation_schema_not_valid_with_other_foreign_key(self):
        """
        only a foreign key to the site code is accepted.
        """
        schema_fields = [
            {
                "name": "What",
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
            {
                "name": "When",
                "type": "date",
                "constraints": helpers.REQUIRED_CONSTRAINTS,
                "biosys": {
                    'type': 'observationDate'
                }
            },
            {
                "name": "Project",  # project not site
                "type": "string",
                "constraints": helpers.REQUIRED_CONSTRAINTS
            },
        ]
        schema = helpers.create_schema_from_fields(schema_fields)
        schema = helpers.add_model_field_foreign_key_to_schema(schema, {
            'schema_field': 'Project',  # project not site
            'model': 'Project',
            'model_field': 'title'
        })
        data_package = helpers.create_data_package_from_schema(schema)
        # create data set
        url = reverse('api:dataset-list')
        project = self.project_1
        client = self.custodian_1_client
        dataset_name = "Observation with project foreign key and no geometry"
        payload = {
            "name": dataset_name,
            "type": Dataset.TYPE_OBSERVATION,
            "project": project.pk,
            'data_package': data_package
        }

        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_geometry_extracted_create(self):
        """
        Test that the record geometry is properly copied from the site when posting through api
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_site_code_fk()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        site_code = 'Cottesloe'
        site_geometry = Point(115.76, -32.0)
        # create the site
        site = G(Site, code=site_code, geometry=site_geometry, project=project)
        record_data = {
            'What': 'Hello! This is a test.',
            'When': '12/12/2017',
            'Site Code': site_code
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.site, site)
        self.assertEqual(record.geometry, site_geometry)

    def test_geometry_extracted_update(self):
        """
        Test that the record geometry is properly copied from the site when updating/patching
        """
        # create the record
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_site_code_fk()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        site_code = 'Cottesloe'
        site_geometry = Point(115.76, -32.0)
        # create the site
        G(Site, code=site_code, geometry=site_geometry, project=project)
        record_data = {
            'What': 'Hello! This is a test.',
            'When': '12/12/2017',
            'Site Code': site_code
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.geometry, site_geometry)

        # update record with new site
        site_code = 'Somewhere'
        site_geometry = Point(116.0, -30.0)
        # create the site
        G(Site, code=site_code, geometry=site_geometry, project=project)
        record_data = {
            'What': 'Yellow!',
            'When': '01/01/2017',
            'Site Code': site_code
        }
        payload = {
            'data': record_data
        }
        url = reverse('api:record-detail', kwargs={'pk': record.pk})
        resp = client.patch(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertIsNotNone(record)
        self.assertEqual(timezone.make_naive(record.datetime), datetime.datetime(2017, 1, 1, 0, 0))
        self.assertEqual(record.geometry, site_geometry)

    def test_record_rejected_if_site_has_no_geometry_api(self):
        """
        When using api
        If the referenced site has no geometry the record should be rejected
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_site_code_fk()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        site_code = 'Cottesloe'
        # create the site
        site = G(Site, code=site_code, geometry=None, project=project)
        self.assertIsNone(site.geometry)
        record_data = {
            'What': 'Hello! This is a test.',
            'When': '12/12/2017',
            'Site Code': site_code
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
        # check error
        errors = resp.json().get('data')
        # errors is of string of format 'field_name::message'
        self.assertIsNotNone(errors)
        self.assertTrue(isinstance(errors, list))
        self.assertEqual(len(errors), 1)
        field_name, message = errors[0].split('::')
        self.assertEqual(field_name, 'Site Code')
        # message should be something like:
        expected_message = 'The site Cottesloe does not exist or has no geometry'
        self.assertEqual(message, expected_message)

    def test_schema_with_lat_long_and_site_fk(self):
        """
        Use case:
        The schema contains a classic lat/lon fields and a site_code foreign key.
        Test that:
        1 - the lat/long provided takes precedence over the site geometry
        2 - if lat/long not provided the site geometry is used
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_latlong_and_site_code_fk()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.latitude_field)
        self.assertIsNotNone(dataset.schema.longitude_field)
        site_code = 'Cottesloe'
        site_geometry = Point(115.76, -32.0)
        # create the site
        site = G(Site, code=site_code, geometry=site_geometry, project=project)
        # the observation geometry different than the site geometry
        observation_geometry = Point(site_geometry.x + 2, site_geometry.y + 2)
        self.assertNotEqual(site.geometry, observation_geometry)

        # lat/long + site
        record_data = {
            'What': 'Hello! This is a test.',
            'When': '12/12/2017',
            'Longitude': observation_geometry.x,
            'Latitude': observation_geometry.y,
            'Site Code': site_code
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.site, site)
        self.assertEqual(record.geometry.geojson, observation_geometry.geojson)

        # lat/long no site
        record_data = {
            'What': 'Hello! This is a test.',
            'When': '12/12/2017',
            'Longitude': observation_geometry.x,
            'Latitude': observation_geometry.y,
            'Site Code': None
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertIsNone(record.site)
        self.assertEqual(record.geometry.geojson, observation_geometry.geojson)

        # site without lat/long
        record_data = {
            'What': 'Hello! This is a test.',
            'When': '12/12/2017',
            'Longitude': None,
            'Latitude': None,
            'Site Code': site_code
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        self.assertEqual(record.site, site)
        self.assertEqual(record.geometry.geojson, site_geometry.geojson)

        # no lat/long no site -> error
        record_data = {
            'What': 'Hello! This is a test.',
            'When': '12/12/2017',
            'Longitude': None,
            'Latitude': None,
            'Site Code': None
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)

    def test_geometry_extracted_upload(self):
        """
        Test that the record geometry is properly copied from the site when using an xlsx upload
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_site_code_fk()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        # create two sites
        site_1_code = 'Cottesloe'
        site_1_geometry = Point(115.76, -32.0)
        site_1 = G(Site, code=site_1_code, geometry=site_1_geometry, project=project)

        site_2_code = 'Somewhere'
        site_2_geometry = Point(116.0, -30.0)
        # create the site
        site_2 = G(Site, code=site_2_code, geometry=site_2_geometry, project=project)

        # data
        csv_data = [
            ['What', 'When', 'Site Code'],
            ['what_1', '01/01/2017', site_1_code],
            ['what_2', '02/02/2017', site_2_code]
        ]
        file_ = helpers.rows_to_xlsx_file(csv_data)
        self.assertEqual(0, Record.objects.filter(dataset=dataset).count())
        url = reverse('api:dataset-upload', kwargs={'pk': dataset.pk})
        with open(file_, 'rb') as fp:
            payload = {
                'file': fp
            }
            resp = client.post(url, data=payload, format='multipart')
            self.assertEqual(status.HTTP_200_OK, resp.status_code)
            records = Record.objects.filter(dataset=dataset)
            self.assertEqual(records.count(), len(csv_data) - 1)
            r = [r for r in records if r.data['What'] == 'what_1'][0]
            self.assertEqual(r.site, site_1)
            self.assertEqual(r.geometry, site_1_geometry)
            r = [r for r in records if r.data['What'] == 'what_2'][0]
            self.assertEqual(r.site, site_2)
            self.assertEqual(r.geometry, site_2_geometry)

    def test_record_rejected_if_site_has_no_geometry_upload(self):
        """
        When uploading with Excel
        If the referenced site has no geometry the record should be rejected
        """
        # same as above but site_2 has no geometry
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_site_code_fk()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        # create two sites the number 2 without a geometry
        site_1_code = 'Cottesloe'
        site_1_geometry = Point(115.76, -32.0)
        site_1 = G(Site, code=site_1_code, geometry=site_1_geometry, project=project)

        site_2_code = 'Somewhere'
        site_2_geometry = None
        G(Site, code=site_2_code, geometry=site_2_geometry, project=project)

        csv_data = [
            ['What', 'When', 'Site Code'],
            ['what_1', '01/01/2017', site_1_code],
            ['what_2', '02/02/2017', site_2_code]
        ]
        file_ = helpers.rows_to_xlsx_file(csv_data)
        self.assertEqual(0, Record.objects.filter(dataset=dataset).count())
        url = reverse('api:dataset-upload', kwargs={'pk': dataset.pk})
        with open(file_, 'rb') as fp:
            payload = {
                'file': fp
            }
            resp = client.post(url, data=payload, format='multipart')
            self.assertEqual(resp.status_code, status.HTTP_400_BAD_REQUEST)
            # Check that the good record is there.
            records = Record.objects.filter(dataset=dataset)
            self.assertEqual(records.count(), 1)
            r = records.first()
            self.assertEqual(r.site, site_1)
            self.assertEqual(r.geometry, site_1_geometry)

    def test_site_geometry_updated(self):
        """
        Use case:
        observations has been created with a site geometry, user update the site location.
        user expect that the associated observations have their geometry updated.
        This can only if the observations has the site as a FK (of course) and exactly the same geometry.
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.schema_with_site_code_fk()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        # create two sites
        site_1_code = 'Cottesloe'
        site_1_geometry = Point(115.76, -32.0)
        site_1 = G(Site, code=site_1_code, geometry=site_1_geometry, project=project)

        site_2_code = 'Somewhere'
        site_2_geometry = Point(116.0, -30.0)
        # create the site
        site_2 = G(Site, code=site_2_code, geometry=site_2_geometry, project=project)

        # data
        csv_data = [
            ['What', 'When', 'Site Code'],
            ['what_1', '01/01/2017', site_1_code],
            ['what_2', '02/02/2017', site_2_code]
        ]
        file_ = helpers.rows_to_xlsx_file(csv_data)
        self.assertEqual(0, Record.objects.filter(dataset=dataset).count())
        url = reverse('api:dataset-upload', kwargs={'pk': dataset.pk})
        with open(file_, 'rb') as fp:
            payload = {
                'file': fp
            }
            resp = client.post(url, data=payload, format='multipart')
            self.assertEqual(status.HTTP_200_OK, resp.status_code)
            records = Record.objects.filter(dataset=dataset)
            self.assertEqual(records.count(), len(csv_data) - 1)
            record_1 = [r for r in records if r.data['What'] == 'what_1'][0]
            self.assertEqual(record_1.site, site_1)
            self.assertEqual(record_1.geometry, site_1_geometry)
            record_2 = [r for r in records if r.data['What'] == 'what_2'][0]
            self.assertEqual(record_2.site, site_2)
            self.assertEqual(record_2.geometry, site_2_geometry)

            # Change the site_1 geometry and expect the record_1 to have its geometry updated
            previous_geometry = site_1_geometry
            new_geometry = Point(previous_geometry.x + 2, previous_geometry.y + 2)
            self.assertNotEqual(previous_geometry, new_geometry)
            url = reverse('api:site-detail', kwargs={'pk': site_1.pk})
            payload = {
                "geometry": new_geometry.wkt
            }
            resp = client.patch(url, data=payload, format='json')
            self.assertEqual(resp.status_code, 200)
            # check that the record has been updated
            record_1.refresh_from_db()
            self.assertEqual(record_1.geometry.geojson, new_geometry.geojson)
            # site_2 record should be untouched
            record_2.refresh_from_db()
            self.assertEqual(record_2.geometry.geojson, site_2_geometry.geojson)

            # Use case: the record geometry should be updated ONLY if it matches exactly the site geometry
            # new geometry for record_1
            new_site_geometry = Point(179, -30)
            self.assertNotEqual(new_site_geometry, site_1.geometry)
            new_record_geometry = Point(180, -35)
            self.assertNotEqual(new_record_geometry, new_site_geometry)
            record_1.geometry = new_record_geometry
            record_1.save()
            self.assertNotEqual(record_1.geometry.geojson, record_1.site.geometry.geojson)
            site = record_1.site
            site.geometry = new_site_geometry
            site.save()
            # check record not changed
            record_1.refresh_from_db()
            self.assertEqual(record_1.geometry, new_record_geometry)
            self.assertNotEqual(record_1.geometry.geojson, record_1.site.geometry.geojson)


class TestMultipleGeometrySource(helpers.BaseUserTestCase):

    def test_geometry_easting_northing_precedence(self):
        """
        If all fields are provided easting and northing have precedence over lat/long and site code.
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.observation_schema_with_with_all_possible_geometry_fields()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.datum_field)
        self.assertIsNotNone(dataset.schema.zone_field)

        # site geometry
        site_code = 'Cottesloe'
        site_geometry = Point(115.76, -32.0)
        # create the site
        site = G(Site, code=site_code, geometry=site_geometry, project=project)

        # easting/northing: nearly (116.0, -32.0)
        easting = 405542.537
        northing = 6459127.469
        east_north_datum = 'GDA94'
        zone = 50
        east_north_srid = 28350

        # lat/long
        longitude = 117.0
        latitude = -33.0
        lat_long_datum = 'WGS84'
        lat_long_srid = 4326

        record_data = {
            'What': 'A record with all geometry fields populated',
            'When': '12/12/2017',
            'Site Code': site_code,
            'Easting': easting,
            'Northing': northing,
            'Datum': east_north_datum,
            'Zone': zone,
            'Latitude': latitude,
            'Longitude': longitude
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        geometry = record.geometry
        self.assertIsNotNone(geometry)
        self.assertIsInstance(geometry, Point)
        # it should be the easting/northing geometry
        geometry.transform(east_north_srid)
        self.assertAlmostEqual(geometry.x, easting, places=2)
        self.assertAlmostEqual(geometry.y, northing, places=2)

    def test_geometry_lat_long_precedence(self):
        """
        Lat/long takes precedence over site code
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.observation_schema_with_with_all_possible_geometry_fields()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.datum_field)
        self.assertIsNotNone(dataset.schema.zone_field)

        # site geometry
        site_code = 'Cottesloe'
        site_geometry = Point(115.76, -32.0)
        # create the site
        site = G(Site, code=site_code, geometry=site_geometry, project=project)

        # lat/long
        longitude = 117.0
        latitude = -33.0
        lat_long_datum = 'WGS84'
        lat_long_srid = 4326

        record_data = {
            'What': 'A record with all geometry fields populated',
            'When': '12/12/2017',
            'Site Code': site_code,
            'Easting': None,
            'Northing': None,
            'Datum': lat_long_datum,
            'Zone': None,
            'Latitude': latitude,
            'Longitude': longitude
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        geometry = record.geometry
        self.assertIsNotNone(geometry)
        self.assertIsInstance(geometry, Point)
        # it should be the lat/long geometry
        geometry.transform(lat_long_srid)
        self.assertAlmostEqual(geometry.x, longitude, places=4)
        self.assertAlmostEqual(geometry.y, latitude, places=4)
        # and not the site
        self.assertNotAlmostEqual(geometry.x, site_geometry.x, places=4)
        self.assertNotAlmostEqual(geometry.y, site_geometry.y, places=4)

    def test_easting_northing_and_site(self):
        """
        Easting/Northing > site_code
        """
        project = self.project_1
        client = self.custodian_1_client
        schema = self.observation_schema_with_with_all_possible_geometry_fields()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.datum_field)
        self.assertIsNotNone(dataset.schema.zone_field)

        # site geometry
        site_code = 'Cottesloe'
        site_geometry = Point(115.76, -32.0)
        # create the site
        site = G(Site, code=site_code, geometry=site_geometry, project=project)

        # easting/northing: nearly (116.0, -32.0)
        easting = 405542.537
        northing = 6459127.469
        east_north_datum = 'GDA94'
        zone = 50
        east_north_srid = 28350

        record_data = {
            'What': 'A record with all geometry fields populated',
            'When': '12/12/2017',
            'Site Code': site_code,
            'Easting': easting,
            'Northing': northing,
            'Datum': east_north_datum,
            'Zone': zone,
            'Latitude': None,
            'Longitude': None
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        geometry = record.geometry
        self.assertIsNotNone(geometry)
        self.assertIsInstance(geometry, Point)
        # it should be the easting/northing geometry
        geometry.transform(east_north_srid)
        self.assertAlmostEqual(geometry.x, easting, places=2)
        self.assertAlmostEqual(geometry.y, northing, places=2)

    def test_site_only(self):
        project = self.project_1
        client = self.custodian_1_client
        schema = self.observation_schema_with_with_all_possible_geometry_fields()
        dataset = self._create_dataset_with_schema(project, client, schema, dataset_type=Dataset.TYPE_OBSERVATION)
        self.assertIsNotNone(dataset.schema.datum_field)
        self.assertIsNotNone(dataset.schema.zone_field)

        # site geometry
        site_code = 'Cottesloe'
        site_geometry = Point(115.76, -32.0)
        # create the site
        site = G(Site, code=site_code, geometry=site_geometry, project=project)

        record_data = {
            'What': 'A record with all geometry fields populated',
            'When': '12/12/2017',
            'Site Code': site_code,
            'Easting': None,
            'Northing': None,
            'Datum': None,
            'Zone': None,
            'Latitude': None,
            'Longitude': None
        }
        payload = {
            'dataset': dataset.pk,
            'data': record_data
        }
        url = reverse('api:record-list')
        resp = client.post(url, data=payload, format='json')
        self.assertEqual(resp.status_code, status.HTTP_201_CREATED)
        record = Record.objects.filter(id=resp.json().get('id')).first()
        self.assertIsNotNone(record)
        geometry = record.geometry
        self.assertIsNotNone(geometry)
        self.assertIsInstance(geometry, Point)
        # it should be the easting/northing geometry
        self.assertAlmostEqual(geometry.x, site_geometry.x, places=4)
        self.assertAlmostEqual(geometry.y, site_geometry.y, places=4)


class TestGeometryConversion(helpers.BaseUserTestCase):

    def test_lat_long_with_projected_project_datum(self):
        """
        see: https://youtrack.gaiaresources.com.au/youtrack/issue/BIOSYS-152
        Use case:
         - Project datum set to be a projected one, e.g GDA/Zone 56.
         - Schema has a latitude, longitude, datum, zone, easting and northing field (the whole shebang)
         - Post a record with lat=-32.0 long=115.75 and Datum=WGS84
        Success if record.geometry is Point(115.75, -32.0)
        """
        # Create project with projected datum
        datum_srid = constants.get_datum_srid('GDA94 / MGA zone 56')
        project = factories.ProjectFactory.create(
            datum=datum_srid
        )
        self.assertTrue(constants.is_projected_srid(project.datum))
        project.custodians.add(self.custodian_1_user)
        self.assertTrue(project.is_custodian(self.custodian_1_user))

        # Dataset and records in lat/long
        schema = self.observation_schema_with_with_all_possible_geometry_fields()
        client = self.custodian_1_client
        dataset = self._create_dataset_with_schema(
            project,
            client,
            schema,
            dataset_type=Dataset.TYPE_OBSERVATION
        )
        # post record
        record_data = {
            'When': "2018-05-25",
            'Datum': 'WGS84',
            'Latitude': -32.0,
            'Longitude': 115.75
        }
        record = self._create_record(
            client,
            dataset,
            record_data
        )
        self.assertIsNotNone(record)
        self.assertIsNotNone(record.geometry)
        self.assertEqual(record.geometry.x, 115.75)
        self.assertEqual(record.geometry.y, -32.0)

        # try with the upload end-point
        dataset.record_set.all().delete()
        rows = [
            ['When', 'Datum', 'Latitude', 'Longitude'],
            ['2018-05-25', 'WGS84', -32.0, 115.75]
        ]
        response = self._upload_records_from_rows(
            rows,
            dataset.pk,
            strict=True
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        record = dataset.record_set.last()
        self.assertIsNotNone(record)
        self.assertIsNotNone(record.geometry)
        self.assertEqual(record.geometry.x, 115.75)
        self.assertEqual(record.geometry.y, -32.0)


class TestSerialization(TestCase):
    fixtures = [
        'test-users',
        'test-projects',
        'test-sites',
        'test-datasets',
        'test-observations'
    ]

    @override_settings(PASSWORD_HASHERS=('django.contrib.auth.hashers.MD5PasswordHasher',),
                       REST_FRAMEWORK_TEST_SETTINGS=helpers.REST_FRAMEWORK_TEST_SETTINGS)
    def setUp(self):
        password = 'password'
        self.admin_user = User.objects.filter(username="admin").first()
        self.assertIsNotNone(self.admin_user)
        self.assertTrue(is_admin(self.admin_user))
        self.admin_user.set_password(password)
        self.admin_user.save()
        self.admin_client = APIClient()
        self.assertTrue(self.admin_client.login(username=self.admin_user.username, password=password))

        self.custodian_1_user = User.objects.filter(username="custodian1").first()
        self.assertIsNotNone(self.custodian_1_user)
        self.custodian_1_user.set_password(password)
        self.custodian_1_user.save()
        self.custodian_1_client = APIClient()
        self.assertTrue(self.custodian_1_client.login(username=self.custodian_1_user.username, password=password))
        self.project_1 = Project.objects.filter(name="Project1").first()
        self.site_1 = Site.objects.filter(code="Adolphus").first()
        self.ds_1 = Dataset.objects.filter(name="Observation1", project=self.project_1).first()
        self.assertIsNotNone(self.ds_1)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = self.ds_1.record_model.objects.filter(dataset=self.ds_1).first()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))
        self.assertTrue(self.record_1.site, self.site_1)

        self.custodian_2_user = User.objects.filter(username="custodian2").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.custodian_2_user.set_password(password)
        self.custodian_2_user.save()
        self.custodian_2_client = APIClient()
        self.assertTrue(self.custodian_2_client.login(username=self.custodian_2_user.username, password=password))
        self.project_2 = Project.objects.filter(name="Project2").first()
        self.site_2 = Site.objects.filter(code="Site2").first()
        self.ds_2 = Dataset.objects.filter(name="Observation2", project=self.project_2).first()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

        self.readonly_user = User.objects.filter(username="readonly").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.assertFalse(self.site_2.is_custodian(self.readonly_user))
        self.assertFalse(self.site_1.is_custodian(self.readonly_user))
        self.readonly_user.set_password(password)
        self.readonly_user.save()
        self.readonly_client = APIClient()
        self.assertTrue(self.readonly_client.login(username=self.readonly_user.username, password=password))

        self.anonymous_client = APIClient()

    def test_date_serialization_uses_project_timezone(self):
        # TODO: implement this
        pass


class TestExport(helpers.BaseUserTestCase):
    fixtures = helpers.BaseUserTestCase.fixtures + [
        'test-sites',
        'test-datasets',
        'test-observations'
    ]

    def _more_setup(self):
        self.ds_1 = Dataset.objects.filter(name="Observation1", project=self.project_1).first()
        self.assertIsNotNone(self.ds_1)
        self.assertTrue(self.ds_1.is_custodian(self.custodian_1_user))
        self.record_1 = Record.objects.filter(dataset=self.ds_1).first()
        self.assertIsNotNone(self.record_1)
        self.assertTrue(self.record_1.is_custodian(self.custodian_1_user))

        self.ds_2 = Dataset.objects.filter(name="Bats2", project=self.project_2).first()
        self.assertTrue(self.ds_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.ds_1.is_custodian(self.custodian_2_user))

    def test_happy_path_no_filter(self):
        client = self.custodian_1_client
        dataset = self.ds_1
        all_records = Record.objects.filter(dataset=dataset)
        self.assertTrue(all_records.count() > 0)
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        # check headers
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=DatasetName_YYYY_MM_DD-HHMMSS.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEqual(ext, '.xlsx')
        filename.startswith(dataset.name)
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named from dataset
        sheet_names = wb.sheetnames
        self.assertEqual(1, len(sheet_names))
        self.assertEqual(dataset.name, sheet_names[0])
        ws = wb[dataset.name]
        rows = list(ws.rows)
        expected_records = Record.objects.filter(dataset=dataset)
        self.assertEqual(len(rows), expected_records.count() + 1)
        headers = [c.value for c in rows[0]]
        schema = dataset.schema
        # all the columns of the schema should be in the excel
        self.assertEqual(schema.headers, headers)

    def test_permission_ok_for_not_custodian(self):
        """Export is a read action. Should be authorised for every logged-in user."""
        client = self.custodian_2_client
        dataset = self.ds_1
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_200_OK)

    def test_permission_denied_if_not_logged_in(self):
        """Must be logged-in."""
        client = self.anonymous_client
        dataset = self.ds_1
        url = reverse('api:record-list')
        query = {
            'dataset__id': dataset.pk,
            'output': 'xlsx'
        }
        try:
            resp = client.get(url, query)
        except Exception as e:
            self.fail("Export should not raise an exception: {}".format(e))
        self.assertEqual(resp.status_code, status.HTTP_401_UNAUTHORIZED)


class TestDateNotMandatory(helpers.BaseUserTestCase):
    date_easting_northing_site_nothing_required_schema = [
        {
            "name": "What",
            "type": "string",
            "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
        },
        {
            "name": "When",
            "type": "date",
            "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
            "format": "any",
            "biosys": {
                'type': 'observationDate'
            }
        },
        {
            "name": 'Site',
            "type": "string",
            "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
            "biosys": {
                'type': 'siteCode'
            }
        },
        {
            "name": "Northing",
            "type": "number",
            "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
            "biosys": {
                "type": "northing"
            }
        },
        {
            "name": "Easting",
            "type": "number",
            "constraints": helpers.NOT_REQUIRED_CONSTRAINTS,
            "biosys": {
                "type": "easting"
            }
        },
        {
            "name": "Datum",
            "type": "string",
            "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
        },
        {
            "name": "Zone",
            "type": "integer",
            "constraints": helpers.NOT_REQUIRED_CONSTRAINTS
        }
    ]

    def test_record_without_date(self):
        project = self.project_1
        client = self.custodian_1_client
        dataset = self._create_dataset_with_schema(
            project,
            client,
            self.date_easting_northing_site_nothing_required_schema,
            Dataset.TYPE_OBSERVATION
        )

        # easting/northing: nearly (116.0, -32.0)
        easting = 405542.537
        northing = 6459127.469
        datum = 'GDA94'
        zone = 50
        east_north_srid = 28350

        record_data = {
            'What': 'Whaaat?',
            'Easting': easting,
            'Northing': northing,
            'Datum': datum,
            'Zone': zone
        }
        record = self._create_record(client, dataset, record_data)
        self.assertIsNone(record.datetime)
        geometry = record.geometry
        self.assertIsNotNone(geometry)
        self.assertIsInstance(geometry, Point)
        geometry.transform(east_north_srid)
        self.assertAlmostEqual(geometry.x, easting, places=2)
        self.assertAlmostEqual(geometry.y, northing, places=2)


class TestPatch(helpers.BaseUserTestCase):

    def test_patch_published(self):
        """
        Test that we can patch just the 'published' flag
        :return:
        """
        rows = [
            ['What', 'When', 'Latitude', 'Longitude', 'Comments'],
            ['Chubby bat', '2018-06-01', -32, 115.75, 'It is huge!']
        ]
        dataset = self._create_dataset_and_records_from_rows(rows)
        self.assertEqual(dataset.type, Dataset.TYPE_OBSERVATION)
        records = dataset.record_set.all()
        record = records.last()
        self.assertIsNotNone(record)
        self.assertFalse(record.published)
        previous_data = json.dumps(record.data)
        # patch
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        payload = {
            'published': True
        }
        resp = client.patch(url, payload)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertTrue(record.published)
        self.assertTrue(json.dumps(record.data), previous_data)

    def test_patch_consumed(self):
        """
        Test that we can patch just the 'consumed' flag
        :return:
        """
        rows = [
            ['What', 'When', 'Latitude', 'Longitude', 'Comments'],
            ['Chubby bat', '2018-06-01', -32, 115.75, 'It is huge!']
        ]
        dataset = self._create_dataset_and_records_from_rows(rows)
        self.assertEqual(dataset.type, Dataset.TYPE_OBSERVATION)
        records = dataset.record_set.all()
        record = records.last()
        self.assertIsNotNone(record)
        self.assertFalse(record.consumed)
        previous_data = json.dumps(record.data)
        # patch
        url = reverse('api:record-detail', kwargs={"pk": record.pk})
        client = self.custodian_1_client
        payload = {
            'consumed': True
        }
        resp = client.patch(url, payload)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        record.refresh_from_db()
        self.assertTrue(record.consumed)
        self.assertTrue(json.dumps(record.data), previous_data)
