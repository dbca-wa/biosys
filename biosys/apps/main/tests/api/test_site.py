import json
import re
from os import path

from openpyxl import load_workbook

from django.contrib.auth.models import User
from django.contrib.gis.geos import GEOSGeometry
from django.core.urlresolvers import reverse
from django.test import TestCase, override_settings
from django.utils import six

from rest_framework import status
from rest_framework.test import APIClient

from main.models import Project, Site
from main.tests.api import helpers
from main.utils_auth import is_admin


class TestPermissions(TestCase):
    """
    Test Permissions
    Get: authenticated
    Update: admin, custodians
    Create: admin, custodians
    Delete: admin, custodians
    """
    fixtures = [
        'test-users',
        'test-projects',
        'test-sites'
    ]

    @override_settings(PASSWORD_HASHERS=('django.contrib.auth.hashers.MD5PasswordHasher',),
                       REST_FRAMEWORK_TEST_SETTINGS=helpers.REST_FRAMEWORK_TEST_SETTINGS)
    def setUp(self):
        password = 'password'
        self.admin_user = User.objects.filter(username="admin").first()
        self.assertIsNotNone(self.admin_user)
        self.assertTrue(is_admin(self.admin_user))
        self.admin_user.set_password(password)
        self.admin_user.save()
        self.admin_client = APIClient()
        self.assertTrue(self.admin_client.login(username=self.admin_user.username, password=password))

        self.custodian_1_user = User.objects.filter(username="custodian1").first()
        self.assertIsNotNone(self.custodian_1_user)
        self.custodian_1_user.set_password(password)
        self.custodian_1_user.save()
        self.custodian_1_client = APIClient()
        self.assertTrue(self.custodian_1_client.login(username=self.custodian_1_user.username, password=password))
        self.project_1 = Project.objects.filter(name="Project1").first()
        self.site_1 = Site.objects.filter(code="Site1").first()
        self.assertTrue(self.site_1.is_custodian(self.custodian_1_user))

        self.custodian_2_user = User.objects.filter(username="custodian2").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.custodian_2_user.set_password(password)
        self.custodian_2_user.save()
        self.custodian_2_client = APIClient()
        self.assertTrue(self.custodian_2_client.login(username=self.custodian_2_user.username, password=password))
        self.project_2 = Project.objects.filter(name="Project2").first()
        self.site_2 = Site.objects.filter(code="Site2").first()
        self.assertTrue(self.site_2.is_custodian(self.custodian_2_user))
        self.assertFalse(self.site_1.is_custodian(self.custodian_2_user))

        self.readonly_user = User.objects.filter(username="readonly").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.assertFalse(self.site_2.is_custodian(self.readonly_user))
        self.assertFalse(self.site_1.is_custodian(self.readonly_user))
        self.readonly_user.set_password(password)
        self.readonly_user.save()
        self.readonly_client = APIClient()
        self.assertTrue(self.readonly_client.login(username=self.readonly_user.username, password=password))

        self.anonymous_client = APIClient()

    def test_get(self):
        urls = [
            reverse('api:site-list'),
            reverse('api:site-detail', kwargs={'pk': 1})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.get(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        # authenticated
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.get(url).status_code,
                    status.HTTP_200_OK
                )

    def test_create(self):
        """
        Admin and custodians
        :return:
        """
        project = self.project_1
        urls = [reverse('api:site-list')]
        data = {
            "name": "A new project for Unit test",
            "code": "C12",
            "project": project.pk
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                # site code must be unique
                data['code'] += '1'
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )

    def test_bulk_create(self):
        """
        Cannot create bulk with this end point
        :return:
        """
        project = self.project_1
        urls = [reverse('api:site-list')]
        data = [
            {
                "name": "A new project for Unit test",
                "code": "C1111",
                "project": project.pk
            },
            {
                "name": "A new project for Unit test",
                "code": "C2222",
                "project": project.pk
            }
        ]
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client, self.admin_client,
                          self.custodian_1_client],
            "allowed": []
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data, format='json').status_code,
                    [status.HTTP_400_BAD_REQUEST, status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                # site code must be unique
                for site in data:
                    site['code'] += '1'
                self.assertEqual(
                    client.post(url, data, format='json').status_code,
                    status.HTTP_201_CREATED
                )

    def test_update1(self):
        """
        admin + custodian of project for site 1
        :return:
        """
        site = self.site_1
        previous_code = site.code
        updated_code = previous_code + "-updated"
        urls = [reverse('api:site-detail', kwargs={'pk': site.pk})]
        data = {
            "code": updated_code,
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.patch(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                site.code = previous_code
                site.save()
                self.assertEqual(
                    client.patch(url, data, format='json').status_code,
                    status.HTTP_200_OK
                )
                site.refresh_from_db()
                self.assertEqual(site.code, updated_code)

    def test_update2(self):
        """
        admin + custodian of project for site 2
        :return:
        """
        site = self.site_2
        previous_code = site.code
        updated_code = previous_code + "-updated"
        urls = [reverse('api:site-detail', kwargs={'pk': site.pk})]
        data = {
            "code": updated_code,
        }
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_1_client],
            "allowed": [self.admin_client, self.custodian_2_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.patch(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                site.code = previous_code
                site.save()
                self.assertEqual(
                    client.patch(url, data, format='json').status_code,
                    status.HTTP_200_OK
                )
                site.refresh_from_db()
                self.assertEqual(site.code, updated_code)

    def test_delete(self):
        """
        Currently admin + custodian
        :return:
        """
        site = self.site_1
        urls = [reverse('api:site-detail', kwargs={'pk': site.pk})]
        data = None
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, self.custodian_1_client]
        }

        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.delete(url, data, format='json').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        for client in access['allowed']:
            for url in urls:
                site.save()
                site_count = Site.objects.count()
                self.assertEqual(
                    client.delete(url, data, format='json').status_code,
                    status.HTTP_204_NO_CONTENT
                )
                self.assertTrue(Site.objects.count(), site_count - 1)

    def test_options(self):
        urls = [
            reverse('api:site-list'),
            reverse('api:site-detail', kwargs={'pk': 1})
        ]
        access = {
            "forbidden": [self.anonymous_client],
            "allowed": [self.readonly_client, self.custodian_1_client, self.custodian_2_client, self.admin_client]
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.options(url).status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )
        # authenticated
        for client in access['allowed']:
            for url in urls:
                self.assertEqual(
                    client.options(url).status_code,
                    status.HTTP_200_OK
                )


class TestSiteUpload(TestCase):
    fixtures = [
        'test-users',
        'test-projects'
    ]

    @override_settings(PASSWORD_HASHERS=('django.contrib.auth.hashers.MD5PasswordHasher',),
                       REST_FRAMEWORK_TEST_SETTINGS=helpers.REST_FRAMEWORK_TEST_SETTINGS)
    def setUp(self):
        password = 'password'
        self.admin_user = User.objects.filter(username="admin").first()
        self.assertIsNotNone(self.admin_user)
        self.assertTrue(is_admin(self.admin_user))
        self.admin_user.set_password(password)
        self.admin_user.save()
        self.admin_client = APIClient()
        self.assertTrue(self.admin_client.login(username=self.admin_user.username, password=password))

        self.custodian_1_user = User.objects.filter(username="custodian1").first()
        self.assertIsNotNone(self.custodian_1_user)
        self.custodian_1_user.set_password(password)
        self.custodian_1_user.save()
        self.custodian_1_client = APIClient()
        self.assertTrue(self.custodian_1_client.login(username=self.custodian_1_user.username, password=password))
        self.project_1 = Project.objects.filter(name="Project1").first()

        self.custodian_2_user = User.objects.filter(username="custodian2").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.custodian_2_user.set_password(password)
        self.custodian_2_user.save()
        self.custodian_2_client = APIClient()
        self.assertTrue(self.custodian_2_client.login(username=self.custodian_2_user.username, password=password))
        self.project_2 = Project.objects.filter(name="Project2").first()

        self.readonly_user = User.objects.filter(username="readonly").first()
        self.assertIsNotNone(self.custodian_2_user)
        self.readonly_user.set_password(password)
        self.readonly_user.save()
        self.readonly_client = APIClient()
        self.assertTrue(self.readonly_client.login(username=self.readonly_user.username, password=password))

        self.anonymous_client = APIClient()

    def test_permissions(self):
        """
        Only custodian or admin
        :return:
        """
        project = self.project_1
        custodian_client = self.custodian_1_client

        urls = [reverse('api:upload-sites', kwargs={'pk': project.pk})]
        access = {
            "forbidden": [self.anonymous_client, self.readonly_client, self.custodian_2_client],
            "allowed": [self.admin_client, custodian_client]
        }
        data = {
            'file': 'dsddsds'
        }
        for client in access['forbidden']:
            for url in urls:
                self.assertIn(
                    client.post(url, data=data, format='multipart').status_code,
                    [status.HTTP_401_UNAUTHORIZED, status.HTTP_403_FORBIDDEN]
                )

        csv_file = helpers.to_csv_file([
            ['Site Code'],
            ['C1']
        ])
        for client in access['allowed']:
            for url in urls:
                with open(csv_file) as fp:
                    self.assertIn(
                        client.post(url, data={'file': fp}, format='multipart').status_code,
                        [status.HTTP_200_OK]
                    )

    def test_upload_csv_happy_path(self):
        csv_data = [
            ['Site Code', 'Site Name', 'Description', 'Latitude', 'Longitude', 'Datum', 'Attribute1', 'Attribute2'],
            ['C1', 'Site 1', 'Description1', -32, 116, '', 'attr11', 'attr12'],
            ['C2', 'Site 2', 'Description2', -31, 117, '', 'attr21', 'attr22']
        ]
        csv_file = helpers.to_csv_file(csv_data)
        project = self.project_1
        client = self.custodian_1_client
        url = reverse('api:upload-sites', kwargs={'pk': project.pk})
        self.assertEquals(0, Site.objects.filter(project=project).count())
        with open(csv_file) as fp:
            data = {
                'file': fp
            }
            resp = client.post(url, data=data, format='multipart')
            self.assertEquals(status.HTTP_200_OK, resp.status_code)
            qs = Site.objects.filter(project=project)
            self.assertEquals(len(csv_data) - 1, qs.count())
            self.assertEquals(['C1', 'C2'], [s.code for s in qs.order_by('code')])
            self.assertEquals(['Site 1', 'Site 2'], [s.name for s in qs.order_by('name')])
            self.assertEquals(['Description1', 'Description2'], [s.description for s in qs.order_by('description')])

            # test geom and attr
            s = qs.filter(code='C1').first()
            self.assertEquals((116, -32), (s.geometry.x, s.geometry.y))
            expected_attributes = {
                'Latitude': '-32',
                'Longitude': '116',
                'Datum': '',
                'Attribute1': 'attr11',
                'Attribute2': 'attr12'}

            self.assertEquals(expected_attributes, s.attributes)

            self.assertEqual(project.site_count, len(csv_data) - 1)

    def test_upload_xlsx_happy_path(self):
        csv_data = [
            ['Site Code', 'Site Name', 'Description', 'Latitude', 'Longitude', 'Datum', 'Attribute1', 'Attribute2'],
            ['C1', 'Site 1', 'Description1', -32, 116, '', 'attr11', 'attr12'],
            ['C2', 'Site 2', 'Description2', -31, 117, '', 'attr21', 'attr22']
        ]
        xlsx_file = helpers.to_xlsx_file(csv_data)
        project = self.project_1
        client = self.custodian_1_client
        url = reverse('api:upload-sites', kwargs={'pk': project.pk})
        self.assertEquals(0, Site.objects.filter(project=project).count())
        with open(xlsx_file, 'rb') as fp:
            data = {
                'file': fp
            }
            resp = client.post(url, data=data, format='multipart')
            self.assertEquals(status.HTTP_200_OK, resp.status_code)
            qs = Site.objects.filter(project=project)
            self.assertEquals(len(csv_data) - 1, qs.count())
            self.assertEquals(['C1', 'C2'], [s.code for s in qs.order_by('code')])
            self.assertEquals(['Site 1', 'Site 2'], [s.name for s in qs.order_by('name')])
            self.assertEquals(['Description1', 'Description2'], [s.description for s in qs.order_by('description')])

            # test geom and attr
            s = qs.filter(code='C1').first()
            self.assertEquals((116, -32), (s.geometry.x, s.geometry.y))
            expected_attributes = {
                'Latitude': '-32',
                'Longitude': '116',
                'Datum': '',
                'Attribute1': 'attr11',
                'Attribute2': 'attr12'}

            self.assertEquals(expected_attributes, s.attributes)

            self.assertEqual(project.site_count, len(csv_data) - 1)


class TestSerialization(helpers.BaseUserTestCase):
    fixtures = [
        'test-users',
        'test-projects',
        'test-sites'
    ]

    def test_centroid(self):
        project = self.project_1
        client = self.custodian_1_client
        site = Site.objects.filter(project=project, geometry__isnull=False).first()
        self.assertIsNotNone(site)
        url = reverse('api:site-detail', kwargs={'pk': site.pk})
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        data = resp.json()
        self.assertTrue('centroid' in data)
        centroid = GEOSGeometry(json.dumps(data['centroid']))
        self.assertEqual(centroid, site.geometry.centroid)


class TestDownloadTemplates(helpers.BaseUserTestCase):

    def test_lat_long(self):
        client = self.custodian_1_client
        url = reverse('api:site-template-lat-long')
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=Sites_template_lat_long.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEquals(ext, '.xlsx')
        self.assertEquals(filename, 'Sites_template_lat_long')
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named 'Sites'
        expected_sheet_name = 'Sites'
        sheet_names = wb.get_sheet_names()
        self.assertEquals(1, len(sheet_names))
        self.assertEquals(sheet_names[0], expected_sheet_name)
        ws = wb.get_sheet_by_name(expected_sheet_name)
        rows = list(ws.rows)
        # only one row
        self.assertEquals(len(rows), 1)
        got_headers = [c.value for c in rows[0]]
        expected_headers = ['Name', 'Code', 'Description', 'Latitude', 'Longitude', 'Datum']
        self.assertEqual(got_headers, expected_headers)

    def test_easting_northing(self):
        client = self.custodian_1_client
        url = reverse('api:site-template-easting-northing')
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=Sites_template_lat_long.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEquals(ext, '.xlsx')
        self.assertEquals(filename, 'Sites_template_easting_northing')
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named 'Sites'
        expected_sheet_name = 'Sites'
        sheet_names = wb.get_sheet_names()
        self.assertEquals(1, len(sheet_names))
        self.assertEquals(sheet_names[0], expected_sheet_name)
        ws = wb.get_sheet_by_name(expected_sheet_name)
        rows = list(ws.rows)
        # only one row
        self.assertEquals(len(rows), 1)
        got_headers = [c.value for c in rows[0]]
        expected_headers = ['Name', 'Code', 'Description', 'Easting', 'Northing', 'Datum', 'Zone']
        self.assertEqual(got_headers, expected_headers)
