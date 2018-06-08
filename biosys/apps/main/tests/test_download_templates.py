import re
from os import path

from openpyxl import load_workbook

from django.test import TestCase
from django.test.client import Client
from django.shortcuts import reverse
from django.utils import six
from rest_framework import status


class TestDownloadSiteTemplates(TestCase):

    def test_lat_long_no_logging(self):
        """
        Test lat-long template download.
        Important: Logging should not be necessary
        """
        client = Client()
        url = reverse('download:site-template-lat-long')
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=Sites_template_lat_long.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEqual(ext, '.xlsx')
        self.assertEqual(filename, 'Sites_template_lat_long')
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named 'Sites'
        expected_sheet_name = 'Sites'
        sheet_names = wb.get_sheet_names()
        self.assertEqual(1, len(sheet_names))
        self.assertEqual(sheet_names[0], expected_sheet_name)
        ws = wb[expected_sheet_name]
        rows = list(ws.rows)
        # only one row
        self.assertEqual(len(rows), 1)
        got_headers = [c.value for c in rows[0]]
        expected_headers = ['Name', 'Code', 'Description', 'Latitude', 'Longitude', 'Datum']
        self.assertEqual(got_headers, expected_headers)

    def test_easting_northing_no_logging(self):
        """
        Test easting-northing template download.
        Important: Logging should not be necessary
        """
        client = Client()
        url = reverse('download:site-template-easting-northing')
        resp = client.get(url)
        self.assertEqual(resp.status_code, status.HTTP_200_OK)
        self.assertEqual(resp.get('content-type'),
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        content_disposition = resp.get('content-disposition')
        # should be something like:
        # 'attachment; filename=Sites_template_lat_long.xlsx
        match = re.match('attachment; filename=(.+)', content_disposition)
        self.assertIsNotNone(match)
        filename, ext = path.splitext(match.group(1))
        self.assertEqual(ext, '.xlsx')
        self.assertEqual(filename, 'Sites_template_easting_northing')
        # read content
        wb = load_workbook(six.BytesIO(resp.content), read_only=True)
        # one datasheet named 'Sites'
        expected_sheet_name = 'Sites'
        sheet_names = wb.get_sheet_names()
        self.assertEqual(1, len(sheet_names))
        self.assertEqual(sheet_names[0], expected_sheet_name)
        ws = wb[expected_sheet_name]
        rows = list(ws.rows)
        # only one row
        self.assertEqual(len(rows), 1)
        got_headers = [c.value for c in rows[0]]
        expected_headers = ['Name', 'Code', 'Description', 'Easting', 'Northing', 'Datum', 'Zone']
        self.assertEqual(got_headers, expected_headers)

