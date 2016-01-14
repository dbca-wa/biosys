import logging
from collections import OrderedDict

from openpyxl.worksheet import SheetProtection
from openpyxl.styles import Font, Alignment
from openpyxl import load_workbook

from main.models import Visit, Site, SiteCharacteristic
from species.models import SpeciesFile
from main import utils as util_model
from upload import utils_openpyxl as util_xls
from upload.validation import DATASHEET_MODELS_MAPPING, DATASHEET_META_MAPPING, get_mapping_for_model

logger = logging.getLogger(__name__)

COLUMN_WIDTH_BUFFER = 3
SPECIES_COLUMN_WIDTH = 34


class SiteVisitDatasheetWriter():
    column_header_font = Font(bold=True)
    sheet_protection = SheetProtection(sheet=True, password=None)
    # This is the arbitrary limit of rows we apply data validation or formulas for model that accept multiple rows.
    # Might needs adjustment
    max_row = 100

    def __init__(self, workbook, visit=None, site=None):
        self.wb = workbook
        self.visit = visit
        self.site = site
        self.file_species = SpeciesFile.objects.last()

    def write(self):
        self._write_meta().protection = self.sheet_protection
        self._write_models()
        self._write_lookups().protection = self.sheet_protection
        if self.file_species:
            self._write_species(self.file_species.path).protection = self.sheet_protection
        if self.site is not None:
            self._populate_site_characteristics()
        return self.wb

    def _write_meta(self):
        mapping = DATASHEET_META_MAPPING
        meta_ws = util_xls.get_sheet(self.wb, mapping.sheet_name) or util_xls.get_sheet(self.wb, 'Sheet')
        # rewrite the sheet name in case of the 'Sheet' selection.
        meta_ws.title = mapping.sheet_name
        if meta_ws is None:
            meta_ws = self.wb.create_sheet(0, mapping.sheet_name)
        # write the visit name
        # populate Meta with visit/site details (very rough)
        col_headers = [util_model.get_field_verbose_name(Visit, 'name'),
                       util_model.get_field_verbose_name(Site, 'site_code')]
        # write column headers
        top_cell = meta_ws.cell(row=mapping.top_left_row, column=mapping.top_left_column)
        writing_direction = mapping.next_col_direction
        util_xls.write_values_from_cell(top_cell, col_headers, writing_direction, self.column_header_font)
        # write values
        if self.visit and self.site:
            values = [self.visit.name, self.site.site_code]
            top_cell = util_xls.get_cell_neighbour(top_cell, mapping.next_row_direction)
            util_xls.write_values_from_cell(top_cell, values, writing_direction)
        return meta_ws

    def _write_models(self):
        for mapping in DATASHEET_MODELS_MAPPING:
            self._write_model(mapping)
        return self.wb

    def _write_model(self, mapping):
        model = mapping.model
        ws = util_xls.get_or_create_sheet(self.wb, mapping.sheet_name)
        fields = util_model.get_datasheet_fields_for_model(model)
        top_cell = ws.cell(row=mapping.top_left_row, column=mapping.top_left_column)
        column_cell = top_cell
        
        # create alignment for transposed header cells
        right_alignment = Alignment(horizontal='right')
        
        max_column_width = 0
        for field in fields:
            # the column header
            col_header = util_model.get_datasheet_field_name(field)
            column_cell.font = self.column_header_font
            column_cell.value = col_header
            if mapping.transpose:
                column_cell.alignment = right_alignment

            # calculate column widths
            if util_model.is_species_observation_field(field):
                # special case for species column width
                ws.column_dimensions[column_cell.column].width = SPECIES_COLUMN_WIDTH
            else:
                column_width = len(column_cell.value) + COLUMN_WIDTH_BUFFER
                if mapping.transpose:
                    if column_width > max_column_width:
                        max_column_width = column_width
                    else:
                        column_width = max_column_width
                ws.column_dimensions[column_cell.column].width = column_width
            
            dv = self._create_data_validation(field)
            if dv is not None:
                # apply data validation to the row cell
                row_cell = util_xls.get_cell_neighbour(column_cell, mapping.next_row_direction)
                dv.add(row_cell)
                if not mapping.unique:
                    # it is a multi row model. We need to extend the data validation to other rows
                    # choose a max row
                    max_row = self.max_row
                    for i in range(0, max_row):
                        row_cell = util_xls.get_cell_neighbour(row_cell, mapping.next_row_direction)
                        dv.add(row_cell)
                ws.add_data_validation(dv)
            column_cell = util_xls.get_cell_neighbour(column_cell, mapping.next_col_direction)

        # insert formulas
        for field_name in mapping.formulas:
            field = mapping.model._meta.get_field_by_name(field_name)[0]
            field_header_cell = util_xls.find_cell_by_value(ws, util_model.get_datasheet_field_name(field))
            field_cell = util_xls.get_cell_neighbour(field_header_cell, mapping.next_row_direction)

            parameter_cells = []
            for param_field_name in mapping.formulas[field_name]['parameters']:
                param_field = mapping.model._meta.get_field_by_name(param_field_name)[0]
                param_header_cell = util_xls.find_cell_by_value(ws, util_model.get_datasheet_field_name(param_field))
                parameter_cells.append(util_xls.get_cell_neighbour(param_header_cell, mapping.next_row_direction))

            field_cell.value = self._create_formula(mapping.formulas[field_name]['formula'], *parameter_cells)

            if not mapping.unique:
                # it is a multi row model. We need to extend the formula to other rows choose a max row
                max_row = self.max_row
                for i in range(0, max_row):
                    field_cell = util_xls.get_cell_neighbour(field_cell, mapping.next_row_direction)
                    parameter_cells = [util_xls.get_cell_neighbour(cell, mapping.next_row_direction) for cell in
                                       parameter_cells]
                    field_cell.value = self._create_formula(mapping.formulas[field_name]['formula'], *parameter_cells)

    def _create_data_validation(self, field):
        allow_blank = not util_model.is_mandatory(field)
        dv = None
        if util_model.is_lookup_field(field) and util_model.has_related_objects(field):
            # here we expect that the lookup has been registered as named range.
            # The name of the lookup is the lookup model name (see _write_lookups)
            strict = util_model.is_strict_lookup_field(field)
            lookup_name = util_model.get_field_lookup_model_name(field)
            dv = util_xls.create_list_validation(lookup_name, strict=strict, allow_blank=allow_blank)
        elif util_model.has_choices(field):
            # Should we also add the choices in the Lookups sheet?
            values = [str(choice[1]) for choice in util_model.get_field_choices(field)]
            strict = True
            dv = util_xls.create_list_validation(values, strict=strict, allow_blank=allow_blank)
        elif util_model.is_boolean_field(field):
            allow_blank = True  # blank is False
            values = ['Y', 'N']
            strict = False
            dv = util_xls.create_list_validation(values, strict=strict, allow_blank=allow_blank)
        # species. Hack! Last minute update. We want species data validation on animals only
        elif util_model.is_species_observation_field(field)\
                and self.file_species is not None \
                and field.model._meta.app_label == 'animals':
            # we expect here that a range call species has been registered (see _write_species)
            strict = False
            dv = util_xls.create_list_validation('species', strict=strict, allow_blank=allow_blank)
        return dv

    def _write_lookups(self):
        """
        Write the lookups as columns in the Lookups spreadsheet.
        For each column we register a named range.
        The name will be the name of the Lookup table (model) NOT the model field name because some fields refer to the
        same lookup tables (ex DisturbanceIndicator and the FeralEvidenceLookup)
        :return:
        """
        # First gather all the lookups tables
        lookups = self._build_lookups_dict()  # key=lookup name  value=lookup values(list)
        # Then for every lookups write a column in Lookup datasheet and register the range
        ws = util_xls.get_or_create_sheet(self.wb, 'Lookups')
        for name, values in lookups.items():
            if len(values) > 0:
                range_ = util_xls.append_column(ws, values)
                self.wb.create_named_range(name, ws, range_)
        return ws

    def _write_species(self, file_path):
        ws = util_xls.get_or_create_sheet(self.wb, 'Species')
        species_ws = load_workbook(file_path).active
        col_max = species_ws.get_highest_row()
        species_list = [cell.value for (cell,) in list(species_ws.get_squared_range(1, 1, 1, col_max))]
        range_ = util_xls.append_column(ws, species_list)
        self.wb.create_named_range('species', ws, range_)
        return ws

    def _build_lookups_dict(self):
        """
        Return a dict with key=lookup name
        and value=lookup values (list)
        :return:
        """
        datasheet_models = [mapping.model for mapping in DATASHEET_MODELS_MAPPING]
        result = OrderedDict()
        for model in datasheet_models:
            lookup_fields = util_model.get_lookup_fields(model)
            for field in lookup_fields:
                values = self._build_lookup_validation_list(field)
                lookup_name = util_model.get_field_lookup_model_name(field)
                if lookup_name not in result:
                    result[lookup_name] = values
        return result

    @staticmethod
    def _create_formula(formula, *cells):
        return formula % tuple([cell.coordinate for cell in cells])

    @staticmethod
    def _build_lookup_validation_list(field, sort=True):
        values = []
        if util_model.is_lookup_field(field):
            # we use the code if there is a code or the value
            for value, code in util_model.get_field_lookups(field):
                if code:
                    # BIOSYS-132: if the code is a string that is an integer add it as integer instead of string
                    # to allow input with keyboard
                    if SiteVisitDatasheetWriter._is_integer(code):
                        values.append(int(code))
                    else:
                        values.append(str(code))
                else:
                    values.append(str(value))
        if sort:
            values.sort()
        return values

    @staticmethod
    def _is_integer(value):
        try:
            int(value)
            return True
        except:
            return False

    def _populate_site_characteristics(self):
        # export the site characteristic data from the Site
        mapping = get_mapping_for_model(SiteCharacteristic)
        if mapping is not None:
            row = [
                self.site.underlaying_geology.value if self.site.underlaying_geology else "",
                self.site.closest_water_distance,
                self.site.closest_water_type.value if self.site.closest_water_type else "",
                self.site.landform_pattern.value if self.site.landform_pattern else "",
                self.site.landform_element.value if self.site.landform_element else "",
                self.site.soil_surface_texture.value if self.site.soil_surface_texture else "",
                self.site.soil_colour,
                self.site.comments,
            ]
            ws = util_xls.get_or_create_sheet(self.wb, mapping.sheet_name)
            top_cell = ws.cell(row=mapping.top_left_row, column=mapping.top_left_column)
            start_cell = util_xls.get_cell_neighbour(top_cell, mapping.next_row_direction)
            writing_direction = mapping.next_col_direction
            util_xls.write_values_from_cell(start_cell, row, writing_direction)
