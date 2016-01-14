import os
import tempfile

from openpyxl import load_workbook, Workbook
from django.shortcuts import get_object_or_404

from main.models import Visit, SiteVisitDataSheetTemplate
from main.utils_zip import zip_dir_to_temp_zip, export_zip
from download.utils import SiteVisitDatasheetWriter
from species.utils import species_names_to_excel


def download_visit_datasheet_view(request, pk):
    visit = get_object_or_404(Visit, pk=pk)
    # lookup for template. Choose the latest one
    # @todo: propose a choice of version?
    template = SiteVisitDataSheetTemplate.objects.last()
    working_dir = tempfile.mkdtemp()
    sites = visit.sites.all()
    for site in sites:
        if template is not None:
            workbook = load_workbook(template.path)
        else:
            workbook = Workbook()
        writer = SiteVisitDatasheetWriter(workbook, visit, site)
        workbook = writer.write()
        file_name = "datasheet_visit-{visit}_site-{site}.xlsx".format(visit=str(visit), site=str(site))
        file_path = os.path.join(working_dir, file_name)
        workbook.save(file_path)
    # zip the all working dir
    zip_file = zip_dir_to_temp_zip(working_dir, delete_after=True)
    zip_name = "datasheets_visit-{visit}".format(visit=str(visit))
    return export_zip(zip_file, zip_name, delete_after=True)